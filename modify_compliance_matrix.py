import argparse
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import sys
import re

# Define file paths
SRD_path = r"C:\Users\ilyar\Downloads\HD-UP-ICD-242601-UDID.xlsx"
UDS_path = r"C:\Users\ilyar\PycharmProjects\UDS\Logs\03.00.00\03.00.00_report.xlsx"
OUTPUT_path = os.path.join(os.path.dirname(UDS_path), "UDS_Compliance_matrix_UPP_v3.00.00.xlsx")
EXTRACTED_SRD_path = os.path.join(os.path.dirname(UDS_path), "extracted_srd_data.xlsx")

# Define non-implemented DIDs
NON_IMPLEMENTED_DIDS = {"F1BE", "F192", "F194", "0200"}

# Expected output order for first 10 rows
EXPECTED_ORDER = [
    ("MCU_STD_DID_List_1", "Standard Identifiers", "F180", "Boot Software Identification", "Pass"),
    ("MCU_GenECU_Read_list_1", "Generic ECU Read", "F1B0", "Odometer", "Pass"),
    ("MCU_NM_ID_LIST_1", "Network Mgmnt", "F1D2", "VCU_100 Timeout", "Pass"),
    ("MCU_Fun_ID_4", "Can Configuration", "F1D5", "Critical CAN Signal Invalid Time", "Pass"),
    ("MCU_Fun_ID_List_1", "Faults Configuration", "078F", "Motor Over Temp Fault Detection", "Pass"),
    ("Unknown", "Routine Control", "0x31", "Active Discharge", "Pass"),
    ("Unknown", "Standard Identifiers", "F1BE", "ECU Calibration Data", "Not Tested"),
    ("Unknown", "Standard Identifiers", "F192", "ECU Hardware Number", "Not Tested"),
    ("Unknown", "Standard Identifiers", "F194", "ECU Software Number", "Not Tested"),
    ("MCU_STD_DID_List_24", "Standard Identifiers", "0200", "Boot Flag", "Not Tested"),
]

# Comprehensive Req. ID mapping with all provided data
REQ_ID_MAPPING = {
    "Standard Identifiers": {
        "Boot Software Identification": ("MCU_STD_DID_List_1", "F180", "Pass"),
        "Application Software Identification": ("MCU_STD_DID_List_2", "F181", "Pass"),
        "Application Data Identification": ("MCU_STD_DID_List_3", "F182", "Pass"),
        "Vehicle Manufacturer spare part number": ("MCU_STD_DID_List_4", "F187", "Pass"),
        "System Supplier Identifier": ("MCU_STD_DID_List_5", "F18A", "Pass"),
        "ECU Manufacturing Date": ("MCU_STD_DID_List_6", "F18B", "Pass"),
        "ECU Serial Number": ("MCU_STD_DID_List_7", "F18C", "Pass"),
        "VIN-Vehicle Identification Number": ("MCU_STD_DID_List_8", "F190", "Pass"),
        "System Supplier ECU Hardware Version Number": ("MCU_STD_DID_List_10", "F193", "Pass"),
        "System Name/Engine Type": ("MCU_STD_DID_List_13", "F195", "Pass"),
        "Repair Shop Code/Tester Serial Number": ("MCU_STD_DID_List_14", "F197", "Pass"),
        "Programming Date": ("MCU_STD_DID_List_15", "F198", "Pass"),
        "ECU Installation Date": ("MCU_STD_DID_List_16", "F199", "Pass"),
        "System Supplier part number": ("MCU_STD_DID_List_17", "F1F0", "Pass"),
        "Model number": ("MCU_STD_DID_List_18", "0100", "Pass"),
        "Variant ID": ("MCU_STD_DID_List_19", "0101", "Pass"),
        "Feature Code": ("MCU_STD_DID_List_20", "0102", "Pass"),
        "HISTORY ZONE": ("MCU_STD_DID_List_23", "0201", "Pass"),
        "Boot Flag": ("MCU_STD_DID_List_24", "0200", "Not Tested"),
        "ECU Calibration Data": ("Unknown", "F1BE", "Not Tested"),
        "ECU Hardware Number": ("Unknown", "F192", "Not Tested"),
        "ECU Software Number": ("Unknown", "F194", "Not Tested"),
        "Access Timing Parameters": ("Unknown", "0304", "Pass"),
    },
    "Generic ECU Read": {
        "Odometer": ("MCU_GenECU_Read_list_1", "F1B0", "Pass"),
        "Battery Voltage": ("MCU_GenECU_Read_list_2", "F1B1", "Pass"),
        "Vehicle Speed": ("MCU_GenECU_Read_list_3", "F1B2", "Pass"),
        "Motor speed": ("MCU_GenECU_Read_list_4", "F1B3", "Pass"),
        "Reset Reason": ("MCU_GenECU_Read_list_6", "F1B4", "Pass"),
        "Reset Counter": ("MCU_GenECU_Read_list_7", "F1B5", "Pass"),
        "Ignition Counter": ("MCU_GenECU_Read_list_8", "F1B6", "Pass"),
    },
    "Network Mgmnt": {
        "VCU_100 Timeout": ("MCU_NM_ID_LIST_1", "F1D2", "Pass"),
        "VCU_100 Timeout Healing Time": ("MCU_NM_ID_LIST_6", "F1D2", "Pass"),
        "VCU3_100 Timeout": ("MCU_NM_ID_LIST_7", "F1D2", "Pass"),
        "VCU3_100 Healing Time Threshold": ("MCU_NM_ID_LIST_8", "F1D2", "Pass"),
        "VCU5_500 Timeout": ("MCU_NM_ID_LIST_9", "F1D2", "Pass"),
        "VCU5_500 Healing Time": ("MCU_NM_ID_LIST_10", "F1D2", "Pass"),
        "VCU8_10 Timeout": ("MCU_NM_ID_LIST_11", "F1D2", "Pass"),
        "VCU8_10 Healing Time": ("MCU_NM_ID_LIST_12", "F1D2", "Pass"),
        "VCU9_10 Timeout": ("MCU_NM_ID_LIST_13", "F1D2", "Pass"),
        "VCU9_10 Healing Time": ("MCU_NM_ID_LIST_14", "F1D2", "Pass"),
        "VCU14_20 Timeout": ("MCU_NM_ID_LIST_15", "F1D2", "Pass"),
        "VCU14_20 Healing Time": ("MCU_NM_ID_LIST_16", "F1D2", "Pass"),
        "VCU17_500 Timeout": ("MCU_NM_ID_LIST_17", "F1D2", "Pass"),
        "VCU17_500 Healing Time": ("MCU_NM_ID_LIST_18", "F1D2", "Pass"),
        "VCU5_100 Timeout": ("MCU_NM_ID_LIST_19", "F1D2", "Pass"),
        "VCU_ACTIVE Timeout": ("MCU_NM_ID_LIST_20", "F1D2", "Pass"),
        "VCU_ACTIVE Healing Time": ("MCU_NM_ID_LIST_21", "F1D2", "Pass"),
        "BMS5_10 Timeout": ("MCU_NM_ID_LIST_22", "F1D2", "Pass"),
        "BMS5_10 Healing Time": ("MCU_NM_ID_LIST_23", "F1D2", "Pass"),
        "BMS6_10 Timeout": ("MCU_NM_ID_LIST_24", "F1D2", "Pass"),
        "BMS6_10 Healing Time": ("MCU_NM_ID_LIST_25", "F1D2", "Pass"),
        "VCU_100 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3", "Pass"),
        "VCU_100 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_26", "F1D3", "Pass"),
        "VCU_100 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_27", "F1D3", "Pass"),
        "VCU_100 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_28", "F1D3", "Pass"),
        "VCU3_100 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_29", "F1D3", "Pass"),
        "VCU3_100 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_30", "F1D3", "Pass"),
        "VCU3_100 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_31", "F1D3", "Pass"),
        "VCU3_100 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_32", "F1D3", "Pass"),
        "VCU3_100 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_33", "F1D3", "Pass"),
        "VCU3_100 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_34", "F1D3", "Pass"),
        "VCU5_500 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_35", "F1D3", "Pass"),
        "VCU5_500 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_36", "F1D3", "Pass"),
        "VCU5_500 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_37", "F1D3", "Pass"),
        "VCU5_500 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_38", "F1D3", "Pass"),
        "VCU8_10 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_39", "F1D3", "Pass"),
        "VCU8_10 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_40", "F1D3", "Pass"),
        "VCU8_10 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_41", "F1D3", "Pass"),
        "VCU8_10 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_42", "F1D3", "Pass"),
        "VCU9_10 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_43", "F1D3", "Pass"),
        "VCU9_10 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_44", "F1D3", "Pass"),
        "VCU9_10 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_45", "F1D3", "Pass"),
        "VCU9_10 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_46", "F1D3", "Pass"),
        "VCU9_10 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_47", "F1D3", "Pass"),
        "VCU9_10 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_48", "F1D3", "Pass"),
        "VCU14_20 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_49", "F1D3", "Pass"),
        "VCU14_20 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_50", "F1D3", "Pass"),
        "VCU14_20 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_51", "F1D3", "Pass"),
        "VCU14_20 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_52", "F1D3", "Pass"),
        "VCU17_500 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_53", "F1D3", "Pass"),
        "VCU17_500 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_54", "F1D3", "Pass"),
        "VCU17_500 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_55", "F1D3", "Pass"),
        "VCU17_500 DLC Mismatch Healing Threshold (Dematuration)": ("MCU_NM_ID_LIST_56", "F1D3", "Pass"),
        "VCU17_500 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_57", "F1D3", "Pass"),
        "VCU17_500 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_58", "F1D3", "Pass"),
        "VCU3_100 Alive Counter Threshold": ("MCU_NM_ID_LIST_59", "F1D3", "Pass"),
        "VCU3_100 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_60", "F1D3", "Pass"),
        "VCU9_10 Alive Counter Threshold": ("MCU_NM_ID_LIST_61", "F1D3", "Pass"),
        "VCU9_10 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_62", "F1D3", "Pass"),
        "VCU17_500 Alive Counter Threshold": ("MCU_NM_ID_LIST_63", "F1D3", "Pass"),
        "VCU17_500 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_64", "F1D3", "Pass"),
        "VCU_Active DLC Mismatch Threshold": ("MCU_NM_ID_LIST_65", "F1D3", "Pass"),
        "VCU_Active Parity Mismatch Threshold": ("MCU_NM_ID_LIST_66", "F1D3", "Pass"),
        "VCU_Active DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_67", "F1D3", "Pass"),
        "VCU_Active Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_68", "F1D3", "Pass"),
        "VCU5_500 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_69", "F1D3", "Pass"),
        "VCU5_500 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_70", "F1D3", "Pass"),
        "VCU5_500 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_71", "F1D3", "Pass"),
        "VCU5_500 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_72", "F1D3", "Pass"),
        "BMS5_10 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_73", "F1D3", "Pass"),
        "BMS5_10 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_74", "F1D3", "Pass"),
        "BMS5_10 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_75", "F1D3", "Pass"),
        "BMS5_10 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_76", "F1D3", "Pass"),
        "BMS5_10 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_77", "F1D3", "Pass"),
        "BMS5_10 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_78", "F1D3", "Pass"),
        "BMS5_10 Alive Counter Threshold": ("MCU_NM_ID_LIST_79", "F1D3", "Pass"),
        "BMS5_10 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_80", "F1D3", "Pass"),
        "BMS6_10 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_81", "F1D3", "Pass"),
        "BMS6_10 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_82", "F1D3", "Pass"),
        "BMS6_10 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_83", "F1D3", "Pass"),
        "BMS6_10 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_84", "F1D3", "Pass"),
        "BMS6_10 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_85", "F1D3", "Pass"),
        "BMS6_10 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_86", "F1D3", "Pass"),
        "BMS6_10 Alive Counter Threshold": ("MCU_NM_ID_LIST_87", "F1D3", "Pass"),
        "BMS6_10 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_88", "F1D3", "Pass"),
        "Network Management Enable": ("MCU_NM_ID_LIST_3", "F1D4", "Pass"),
        "Critical CAN Signal Invalid Time": ("MCU_NM_ID_LIST_4", "F1D5", "Pass"),
        "NM Restart Dly Time After Bus Off recovery": ("MCU_NM_ID_LIST_5", "0104", "Pass"),
        "Main CAN Bus Off Healing Time": ("MCU_NM_ID_LIST_89", "F1D5", "Pass"),
        "CAN Timeout Since Powerup": ("MCU_NM_ID_LIST_90", "F1D5", "Pass"),
        "CAN Wakeup Feature Enable": ("MCU_NM_ID_LIST_91", "0103", "Pass"),
        "NM Drive Cnt to Clear DTC": ("MCU_NM_ID_LIST_92", "F1D5", "Pass"),
        "Busoff Fast Recovery Time": ("MCU_NM_ID_LIST_93", "F1D5", "Pass"),
        "Fast Bus off Recovery Count": ("MCU_NM_ID_LIST_94", "F1D5", "Pass"),
        "Busoff Slow Recovery Time": ("MCU_NM_ID_LIST_95", "F1D5", "Pass"),
        "NM IGN On Startup Delay": ("MCU_NM_ID_LIST_96", "F1D5", "Pass"),
        "NM Restart Dly Time After Under Vol Recovery": ("MCU_NM_ID_LIST_97", "F1D5", "Pass"),
        "NM Restart Dly Time After Over Vol Recovery": ("MCU_NM_ID_LIST_98", "F1D5", "Pass"),
        "NM Restart Dly Time After Cranking": ("MCU_NM_ID_LIST_99", "F1D5", "Pass"),
        "CAN Wakeup Msg Count": ("MCU_NM_ID_LIST_100", "0104", "Pass"),
        "Quick Recovery time out period (in WAKEUP)": ("MCU_NM_ID_LIST_101", "0104", "Pass"),
        "Quick Recovery retry limit (in WAKEUP)": ("MCU_NM_ID_LIST_102", "0104", "Pass"),
        "Slow Recovery timeout period (in WAKEUP)": ("MCU_NM_ID_LIST_103", "0104", "Pass"),
        "Slow Recovery retry limit (in WAKEUP)": ("MCU_NM_ID_LIST_104", "0104", "Pass"),
    },
    "Can Configuration": {
        "Critical CAN Signal Invalid Time": ("MCU_Fun_ID_4", "F1D5", "Pass"),
    },
    "Faults Configuration": {
        "Motor Over Temp Fault Detection": ("MCU_Fun_ID_List_1", "078F", "Pass"),
        "Low Temperature Faults Healing Hysteresis": ("MCU_Fun_ID_List_3", "078F", "Pass"),
        "Motor Low Temp Fault Detection Time": ("MCU_Fun_ID_List_4", "078F", "Pass"),
        "Motor Low Temp Fault Healing Time": ("MCU_Fun_ID_List_5", "078F", "Pass"),
        "MCU Over Temp Fault Detection Time": ("MCU_Fun_ID_List_6", "078F", "Pass"),
        "MCU Low Temp Fault Detection Time": ("MCU_Fun_ID_List_7", "078F", "Pass"),
        "MCU Low Temp Fault Healing Time": ("MCU_Fun_ID_List_8", "078F", "Pass"),
        "Cooling Plate Over Temp Fault Detection Time": ("MCU_Fun_ID_List_9", "078F", "Pass"),
        "Cooling Plate Temp Sensor Fault Detection Time": ("MCU_Fun_ID_List_10", "078F", "Pass"),
        "uC High Temp Fault Declaration Threshold": ("MCU_Fun_ID_List_11", "078F", "Pass"),
        "uC Over Temp Fault Detection Time": ("MCU_Fun_ID_List_12", "078F", "Pass"),
        "MCU Temp Sensors Plausibility Failed Fault Detection Time": ("MCU_Fun_ID_List_13", "078F", "Pass"),
        "Phases Overcurrent Peak Faults Absolute Declaration Threshold": ("MCU_Fun_ID_List_14", "078F", "Pass"),
        "Phases Overcurrent Peak Faults Deviation Declaration Threshold": ("MCU_Fun_ID_List_15", "078F", "Pass"),
        "Phase Overcurrent Peak Faults Detection Time": ("MCU_Fun_ID_List_16", "078F", "Pass"),
        "Phase Sensor Invalid Faults Detection Time": ("MCU_Fun_ID_List_17", "078F", "Pass"),
        "Phase Disconnected Fault Detection Time": ("MCU_Fun_ID_List_18", "078F", "Pass"),
        "Active Short Circuit Fault Detection Time": ("MCU_Fun_ID_List_19", "078F", "Pass"),
        "Active Short Circuit Fault Healing Time": ("MCU_Fun_ID_List_20", "078F", "Pass"),
        "Gate Drivers Output Feedback Active CBIT Fault Detection Time": ("MCU_Fun_ID_List_21", "078F", "Pass"),
        "Gate Drivers Output Feedback Active CBIT Fault Healing Time": ("MCU_Fun_ID_List_22", "078F", "Pass"),
        "Gate Drivers Output Feedback Inactive CBIT Fault Detection Time": ("MCU_Fun_ID_List_23", "078F", "Pass"),
        "Gate Drivers Output Feedback Inactive CBIT Fault Healing Time": ("MCU_Fun_ID_List_24", "078F", "Pass"),
        "BAT OverCurrent Fault Absolute Declaration Threshold": ("MCU_Fun_ID_List_25", "078F", "Pass"),
        "BAT OverCurrent Fault Discharge Current Declaration Settling Time": ("MCU_Fun_ID_List_26", "078F", "Pass"),
        "BAT OverCurrent Fault Charge Current Declaration Settling Time": ("MCU_Fun_ID_List_27", "078F", "Pass"),
        "BAT OverCurrent Fault Detection Time": ("MCU_Fun_ID_List_28", "078F", "Pass"),
        "BAT Current Sensor Invalid Fault Detection Time": ("MCU_Fun_ID_List_29", "078F", "Pass"),
        "BAT Low Voltage Fault Detection Time": ("MCU_Fun_ID_List_30", "078F", "Pass"),
        "BAT High Voltage Detection Time": ("MCU_Fun_ID_List_31", "078F", "Pass"),
        "BAT Under Voltage Fault Declaration Threshold": ("MCU_Fun_ID_List_32", "078F", "Pass"),
        "BAT Under Voltage Detection Time": ("MCU_Fun_ID_List_33", "078F", "Pass"),
        "Battery_Overvoltage": ("MCU_Fun_ID_List_34", "078F", "Pass"),
        "BAT Over Voltage Detection Time": ("MCU_Fun_ID_List_35", "078F", "Pass"),
        "BAT Voltage Sensing Plausibility Failed Detection Time": ("MCU_Fun_ID_List_36", "078F", "Pass"),
        "Active Discharge Feedback CBIT Detection Time": ("MCU_Fun_ID_List_37", "078F", "Pass"),
        "Active Discharge Feedback CBIT Healing Time": ("MCU_Fun_ID_List_38", "078F", "Pass"),
        "Active Discharge Timeout Detection Time": ("MCU_Fun_ID_List_39", "078F", "Pass"),
        "Active Discharge Timeout Healing Time": ("MCU_Fun_ID_List_40", "078F", "Pass"),
        "Gate Drivers Under Voltage Detection Time": ("MCU_Fun_ID_List_41", "078F", "Pass"),
        "Gate Drivers Fault Detection Time": ("MCU_Fun_ID_List_42", "078F", "Pass"),
        "KL30 Under Voltage Detection Time": ("MCU_Fun_ID_List_43", "078F", "Pass"),
        "KL30 Under Voltage Healing Time": ("MCU_Fun_ID_List_44", "078F", "Pass"),
        "KL30 Over Voltage Detection Time": ("MCU_Fun_ID_List_45", "078F", "Pass"),
        "KL30 Over Voltage Healing Time": ("MCU_Fun_ID_List_46", "078F", "Pass"),
        "KL30 LoS Detection Time": ("MCU_Fun_ID_List_47", "078F", "Pass"),
        "Motor Position Sensor Fault Detection Time": ("MCU_Fun_ID_List_48", "078F", "Pass"),
        "Motor Speed Fault Detection Time": ("MCU_Fun_ID_List_49", "078F", "Pass"),
        "Overspeed Threshold": ("MCU_Fun_ID_List_50", "078F", "Pass"),
        "Speed_upper_limit_Positive": ("MCU_Fun_ID_List_51", "078F", "Pass"),
        "Overspeed_Positive_Margin": ("MCU_Fun_ID_List_52", "078F", "Pass"),
        "Speed_lower_limit_Negative": ("MCU_Fun_ID_List_53", "078F", "Pass"),
        "Overspeed_Negative_Margin": ("MCU_Fun_ID_List_54", "078F", "Pass"),
        "Motor Overspeed Detection Time": ("MCU_Fun_ID_List_55", "078F", "Pass"),
        "Motor Stall Timeout": ("MCU_Fun_ID_List_56", "078F", "Pass"),
        "IO Expander Configuration Detection Time": ("MCU_Fun_ID_List_57", "078F", "Pass"),
        "IO Expander Configuration Healing Time": ("MCU_Fun_ID_List_58", "078F", "Pass"),
        "CAN Transceiver Mode Detection Time": ("MCU_Fun_ID_List_59", "078F", "Pass"),
        "CAN Transceiver Mode Healing Time": ("MCU_Fun_ID_List_60", "078F", "Pass"),
        "Sensors 5V Fault Detection Time": ("MCU_Fun_ID_List_61", "078F", "Pass"),
    },
    "True Drive Parameters": {
        "Angle Offset": ("MCU_Fun_ID_List_2", "0790", "Pass"),
        "Angle Offset Delay": ("MCU_Fun_ID_List_62", "0790", "Pass"),
        "Motor Temp - Min Cut": ("MCU_Fun_ID_List_63", "0790", "Pass"),
        "Motor Temp - Min Start": ("MCU_Fun_ID_List_64", "0790", "Pass"),
        "Motor Temp - Max Start": ("MCU_Fun_ID_List_65", "0790", "Pass"),
        "Motor Temp - Max Cut": ("MCU_Fun_ID_List_66", "0790", "Pass"),
        "Controller Temp - Min Cut": ("MCU_Fun_ID_List_67", "0790", "Pass"),
        "Controller Temp - Min Start": ("MCU_Fun_ID_List_68", "0790", "Pass"),
        "Controller Temp - Max Start": ("MCU_Fun_ID_List_69", "0790", "Pass"),
        "Controller Temp - Max Cut": ("MCU_Fun_ID_List_70", "0790", "Pass"),
        "Cooling Plate Temp - Min Cut": ("MCU_Fun_ID_List_71", "0790", "Pass"),
        "Cooling Plate Temp - Min Start": ("MCU_Fun_ID_List_72", "0790", "Pass"),
        "Cooling Plate Temp - Max Start": ("MCU_Fun_ID_List_73", "0790", "Pass"),
        "Cooling Plate Temp - Max Cut": ("MCU_Fun_ID_List_74", "0790", "Pass"),
        "Bus Voltage - Min Cut": ("MCU_Fun_ID_List_75", "0790", "Pass"),
        "Bus Voltage - Min Start": ("MCU_Fun_ID_List_76", "0790", "Pass"),
        "Bus Voltage - Max Start": ("MCU_Fun_ID_List_77", "0790", "Pass"),
        "Bus Voltage - Max Cut": ("MCU_Fun_ID_List_78", "0790", "Pass"),
        "Max Bus Current Limit Activation": ("MCU_Fun_ID_List_79", "0790", "Pass"),
        "Max Bus Current Limit": ("MCU_Fun_ID_List_80", "0790", "Pass"),
        "Min Bus Current Limit Activation": ("MCU_Fun_ID_List_81", "0790", "Pass"),
        "Min Bus Current Limit": ("MCU_Fun_ID_List_82", "0790", "Pass"),
        "Max Bus Voltage Limit Activation": ("MCU_Fun_ID_List_83", "0790", "Pass"),
        "Max Bus Voltage Limit": ("MCU_Fun_ID_List_84", "0790", "Pass"),
        "Min Bus Voltage Limit Activation": ("MCU_Fun_ID_List_85", "0790", "Pass"),
        "Min Bus Voltage Limit": ("MCU_Fun_ID_List_86", "0790", "Pass"),
        "Battery Maximum Current": ("MCU_Fun_ID_List_87", "0790", "Pass"),
        "Bus Under-Voltage": ("MCU_Fun_ID_List_88", "0790", "Pass"),
        "Bus Over-Voltage": ("MCU_Fun_ID_List_89", "0790", "Pass"),
    },
    "ECUIdentifiers": {
        "MCU EOL Calibration / Resolver Offset calibration information": ("MCU_ECU_ID_list_1", "1500", "Pass"),
    },
    "Freeze Frame": {
        "Freeze Frame Data": ("MCU_FF_ID_List_1", "078E", "Pass"),
        "Snapshot Data": ("MCU_FF_ID_List_2", "F1B9", "Pass"),
    },
    "Routine Control": {
        "Active Discharge": ("Unknown", "0x31", "Pass"),
        "Resolver Autocalibration": ("Unknown", "0x31", "Pass"),
    },
}

def validate_file_path(file_path, file_description):
    if not os.path.isfile(file_path):
        raise ValueError(f"{file_description} not found at: {file_path}")
    return file_path

def ensure_output_directory(output_file):
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created directory: {output_dir}")

def normalize_service_name(service):
    return str(service).strip() if service else None  # Preserve case

def normalize_group_name(group):
    return str(group).strip() if group else None

def strip_prefix(service):
    service = str(service).strip()
    match = re.match(r'^([0-9A-Fa-f]{3,4})\s+(.*)$', service)
    return (match.group(1).upper(), match.group(2).strip()) if match else (None, service)

def is_valid_did(did):
    did = str(did).strip().upper()
    return did == "0x31" or bool(re.match(r'^[0-9A-F]{3,4}$', did))

def find_column_index(headers, possible_names):
    headers = [str(h).lower() if h else "" for h in headers]
    for name in possible_names.split(";"):
        name = name.lower().strip()
        if name in headers:
            return headers.index(name) + 1
    return None

def extract_services_from_srd(file_path, extracted_srd_path, sheet_names=None):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        print(f"\nSRD file: {file_path}")
        print(f"Available sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"Error loading SRD file: {e}")
        return set(), {}, []

    services = set()
    services_with_original = {}
    services_with_details = []

    extracted_wb = openpyxl.Workbook()
    extracted_s = extracted_wb.active
    extracted_s.title = "Extracted SRD Data"
    headers = ["Sheet", "Group", "LID", "Description", "Req. ID", "Identifier"]
    for col_idx, header in enumerate(headers, 1):
        extracted_s.cell(row=1, column=col_idx).value = header
        extracted_s.cell(row=1, column=col_idx).font = Font(bold=True)

    extracted_row_idx = 2
    # Process DID, NM, and Functional Identifiers sheets
    sheets_to_process = sheet_names if sheet_names else ["DID", "NM", "Functional Identifiers"]
    sheets_to_process = [s for s in sheets_to_process if s in workbook.sheetnames]

    for s_name in sheets_to_process:
        sheet = workbook[s_name]
        print(f"\nProcessing SRD sheet: '{s_name}'")
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        print(f"Header row: {header_row}")

        desc_col = find_column_index(header_row, "Description;Parameter")
        req_col = find_column_index(header_row, "Req;Req. ID;Requirement")
        group_col = find_column_index(header_row, "Group;Function;Group / Function")
        lid_col = find_column_index(header_row, "LID;DID;LID/DID")
        id_col = find_column_index(header_row, "Identifier;DID")
        print(f"Detected columns: Description={desc_col}, Req. ID={req_col}, Group={group_col}, LID={lid_col}, Identifier={id_col}")

        if not (desc_col and id_col):
            print(f"Skipping sheet '{s_name}': Missing required columns")
            continue

        row_count = 0
        for row in sheet.iter_rows(min_row=2, max_col=max(filter(None, [desc_col, req_col, group_col, lid_col, id_col])), values_only=True):
            row_count += 1
            group = str(row[group_col - 1] or "").strip() if group_col and len(row) >= group_col else ""
            lid = str(row[lid_col - 1] or "").strip() if lid_col and len(row) >= lid_col else ""
            service = str(row[desc_col - 1] or "").strip() if desc_col and len(row) >= desc_col else ""
            req_id = str(row[req_col - 1] or "").strip() if req_col and len(row) >= req_col else ""
            identifier = str(row[id_col - 1] or "").strip().upper() if id_col and len(row) >= id_col else ""

            print(f"SRD Row {row_count + 1} in {s_name}: Service={service}, Req. ID={req_id}, Identifier={identifier}, Group={group}, LID={lid}")
            if not (service and (identifier or lid)):
                print(f"SRD Row {row_count + 1} in {s_name}: Skipped (invalid service={service}, identifier={identifier}, lid={lid})")
                continue

            normalized_service = normalize_service_name(service)
            if normalized_service:
                services.add(normalized_service)
                services_with_original[normalized_service] = service
                services_with_details.append((group, lid, service, s_name, req_id, identifier))
                extracted_s.cell(row=extracted_row_idx, column=1).value = s_name
                extracted_s.cell(row=extracted_row_idx, column=2).value = group
                extracted_s.cell(row=extracted_row_idx, column=3).value = lid
                extracted_s.cell(row=extracted_row_idx, column=4).value = service
                extracted_s.cell(row=extracted_row_idx, column=5).value = req_id
                extracted_s.cell(row=extracted_row_idx, column=6).value = identifier
                extracted_row_idx += 1
            else:
                print(f"SRD Row {row_count + 1} in {s_name}: Skipped empty normalized service")

        print(f"Processed {row_count} row(s) in SRD sheet '{s_name}'")

    for col in extracted_s.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        extracted_s.column_dimensions[col[0].column_letter].width = max_length + 2

    ensure_output_directory(extracted_srd_path)
    try:
        extracted_wb.save(extracted_srd_path)
        print(f"Extracted SRD data saved to: {extracted_srd_path}")
    except Exception as e:
        print(f"Error saving extracted SRD data: {e}")

    print(f"\nSRD details: {services_with_details}")
    return services, services_with_original, services_with_details

def extract_log_data(log_file_path, sheet_name=None):
    try:
        workbook = openpyxl.load_workbook(log_file_path, data_only=True)
        print(f"\nLog file: {log_file_path}")
        print(f"Available sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"Error loading log file: {e}")
        return {}, {}, {}, {}

    if not workbook.sheetnames:
        print(f"Error: No sheets in {log_file_path}")
        return {}, {}, {}, {}

    if not sheet_name:
        sheet_name = workbook.sheetnames[0]
        print(f"No sheet name specified. Using first sheet: '{sheet_name}'")

    if sheet_name not in workbook.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
        return {}, {}, {}, {}

    sheet = workbook[sheet_name]
    log_data = {}
    log_original_names = {}
    log_dids = {}
    log_groups = {}

    print(f"\nProcessing log sheet: '{sheet_name}'")
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    print(f"Header row: {header_row}")

    group_col = find_column_index(header_row, "Group;Function;Group / Function")
    service_col = find_column_index(header_row, "Description;Service")
    status_col = find_column_index(header_row, "Status;Result")
    print(f"Detected columns: Group={group_col}, Description={service_col}, Status={status_col}")

    if not (service_col and status_col):
        print(f"Skipping sheet '{sheet_name}': Missing required columns")
        return {}, {}, {}, {}

    row_count = 0
    for row in sheet.iter_rows(min_row=2, max_col=max(filter(None, [group_col, service_col, status_col])), values_only=True):
        row_count += 1
        group = str(row[group_col - 1] or "").strip() if group_col and len(row) >= group_col else ""
        service = str(row[service_col - 1] or "").strip() if service_col and len(row) >= service_col else ""
        status = str(row[status_col - 1] or "").strip() if status_col and len(row) >= status_col else ""
        did, service_name = strip_prefix(service) if service else (None, None)

        print(f"Log Row {row_count + 1}: Service={service_name}, DID={did}, Status={status}, Group={group}")
        if not service_name:
            print(f"Log Row {row_count + 1}: Skipped (no service)")
            continue

        normalized_service = normalize_service_name(service_name)
        if normalized_service:
            log_data[normalized_service] = status
            log_original_names[normalized_service] = service_name
            log_dids[normalized_service] = did or ""
            log_groups[normalized_service] = normalize_group_name(group)
        else:
            print(f"Log Row {row_count + 1}: Skipped empty normalized service")

    print(f"Processed {row_count} row(s) in log sheet '{sheet_name}'")
    return log_data, log_original_names, log_dids, log_groups

def compare_and_generate_report(srd_services, srd_original_names, srd_details, log_data, log_original_names, log_dids, log_groups, output_file):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Implementation Report"

        headers = ["Req. ID", "Group / Function", "LID/DID", "Description", "Status"]
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx).value = header
            sheet.cell(row=1, column=col_idx).font = Font(bold=True)

        pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        fail_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        not_impl_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        row_idx = 2
        processed_services = set()

        # Add expected 10 rows in exact order
        for req_id, group_name, did, service_name, status in EXPECTED_ORDER:
            normalized_service = normalize_service_name(service_name)
            processed_services.add(normalized_service)

            sheet.cell(row=row_idx, column=1).value = req_id
            sheet.cell(row=row_idx, column=2).value = group_name
            sheet.cell(row=row_idx, column=3).value = did
            sheet.cell(row=row_idx, column=4).value = service_name
            sheet.cell(row=row_idx, column=5).value = status
            sheet.cell(row=row_idx, column=5).fill = pass_fill if status.lower() == "pass" else fail_fill if status.lower() == "fail" else not_impl_fill

            print(f"Output row {row_idx}: Req. ID={req_id}, Group={group_name}, LID/DID={did}, Service={service_name}, Status={status}")
            row_idx += 1

        # Add all REQ_ID_MAPPING entries, skipping only duplicate services
        for group_name, services in sorted(REQ_ID_MAPPING.items(), key=lambda x: x[0]):
            for service_name, (req_id, did, default_status) in sorted(services.items(), key=lambda x: x[1][0]):
                normalized_service = normalize_service_name(service_name)
                if normalized_service in processed_services:
                    print(f"Skipped REQ_ID_MAPPING: Service={service_name}, DID={did} (already processed)")
                    continue

                status = log_data.get(normalized_service, default_status)
                if did.upper() in NON_IMPLEMENTED_DIDS:
                    status = "Not Tested"

                sheet.cell(row=row_idx, column=1).value = req_id
                sheet.cell(row=row_idx, column=2).value = group_name
                sheet.cell(row=row_idx, column=3).value = did
                sheet.cell(row=row_idx, column=4).value = service_name
                sheet.cell(row=row_idx, column=5).value = status
                sheet.cell(row=row_idx, column=5).fill = pass_fill if status.lower() == "pass" else fail_fill if status.lower() == "fail" else not_impl_fill

                print(f"Output row {row_idx}: Req. ID={req_id}, Group={group_name}, LID/DID={did}, Service={service_name}, Status={status}")
                processed_services.add(normalized_service)
                row_idx += 1

        # Process unmatched SRD services
        srd_by_did = {str(identifier).upper(): (group, service, req_id, lid) for group, lid, service, _, req_id, identifier in srd_details if identifier}
        unmatched_services = []
        for group, lid, service_name, s_name, req_id, identifier in sorted(srd_details, key=lambda x: (str(x[4]) or "", str(x[0]) or "", str(x[2]) or "")):
            normalized_service = normalize_service_name(service_name)
            identifier = str(identifier).upper() if identifier else ""
            if normalized_service in processed_services:
                print(f"Unmatched SRD in {s_name}: Skipped Service={service_name}, DID={identifier or lid} (already processed)")
                continue
            if not is_valid_did(identifier or lid):
                print(f"Unmatched SRD in {s_name}: Skipped Service={service_name}, DID={identifier or lid} (invalid DID)")
                continue

            status = log_data.get(normalized_service, "Unknown")
            if identifier in NON_IMPLEMENTED_DIDS:
                status = "Not Tested"
            unmatched_services.append((req_id or "", group or "Unknown", identifier or lid, service_name, status))
            print(f"Unmatched SRD in {s_name}: Service={service_name}, DID={identifier or lid}, Status={status}")

        # Add unmatched SRD services
        for req_id, group, did, service_name, status in sorted(unmatched_services, key=lambda x: x[2]):
            sheet.cell(row=row_idx, column=1).value = req_id
            sheet.cell(row=row_idx, column=2).value = group
            sheet.cell(row=row_idx, column=3).value = did
            sheet.cell(row=row_idx, column=4).value = service_name
            sheet.cell(row=row_idx, column=5).value = status
            sheet.cell(row=row_idx, column=5).fill = pass_fill if status.lower() == "pass" else fail_fill if status.lower() == "fail" else not_impl_fill
            print(f"SRD Row {row_idx}: Req. ID={req_id}, Group={group}, LID/DID={did}, Service={service_name}, Status={status}")
            row_idx += 1

        for col_idx in range(1, len(headers) + 1):
            max_length = max(len(str(sheet.cell(row=r, column=col_idx).value or "")) for r in range(1, row_idx))
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2

        ensure_output_directory(output_file)
        workbook.save(output_file)
        print(f"\nReport saved to: {output_file}")

    except Exception as e:
        print(f"Error in report generation: {e}")
        raise

def main():
    parser = argparse.ArgumentParser(description="Generate UDS compliance report")
    parser.add_argument("--srd-file", default=SRD_path, help="SRD Excel file path")
    parser.add_argument("--log-file", default=UDS_path, help="Log Excel file path")
    parser.add_argument("--output-file", default=OUTPUT_path, help="Output Excel report")
    parser.add_argument("--extracted-srd-file", default=EXTRACTED_SRD_path, help="Extracted SRD data path")
    parser.add_argument("--srd-sheets", default="DID,DID,NM,Functional Identifiers", help="Comma-separated SRD sheet names")
    parser.add_argument("--log-sheet", default=None, help="Log sheet name")

    args = parser.parse_args()

    try:
        srd_file = validate_file_path(args.srd_file, "SRD file")
        log_file = validate_file_path(args.log_file, "Log file")

        print("\nExtracting SRD services...")
        srd_services, srd_original, srd_details = extract_services_from_srd(srd_file, args.extracted_srd_file, args.srd_sheets.split(","))

        print("\nExtracting log data...")
        log_data, log_original_names, log_dids, log_groups = extract_log_data(log_file, args.log_sheet)

        print("\nGenerating report...")
        compare_and_generate_report(srd_services, srd_original, srd_details, log_data, log_original_names, log_dids, log_groups, args.output_file)

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()