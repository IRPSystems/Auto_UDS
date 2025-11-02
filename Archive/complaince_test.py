import argparse
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import sys
import re
from Project.UPP import output_with_raw
import subprocess

# Define file paths
SRD_path = r"C:\Users\ilyar\Downloads\HD-UP-ICD-242601-UDID.xlsx"
UDS_path = r"C:\Users\ilyar\PycharmProjects\UDS\Logs\03.01.00\03.01.00_report.xlsx"
OUTPUT_path = os.path.join(os.path.dirname(UDS_path), "UDS_Compliance_matrix_UPP_v3.00.00.xlsx")
EXTRACTED_SRD_path = os.path.join(os.path.dirname(UDS_path), "extracted_srd_data.xlsx")

# Define non-implemented DIDs
NON_IMPLEMENTED_DIDS = {"F1BE", "F192", "F194"}

# Expected output order for first 10 rows
EXPECTED_ORDER = [
    ("MCU_STD_DID_List_1", "Standard Identifiers", "F180", "Boot Software Identification"),
    ("MCU_GenECU_Read_list_1", "Generic ECU Read", "F1B0", "Odometer"),
    ("MCU_NM_ID_LIST_1", "Network Mgmnt", "F1D2", "VCU3_100 Timeout"),
    ("MCU_NM_ID_LIST_4", "Can Configuration", "F1D5", "Critical CAN Signal Invalid Time"),
    ("MCU_Fun_ID_List_1", "Faults Configuration", "078F", "Motor Over Temp Fault Detection"),
    ("Unknown", "Routine Control", "0x31", "Active Discharge"),
    ("MCU_NM_ID_LIST_4", "Standard Identifiers", "F1BE", "Engine State"),
    ("MCU_NM_ID_LIST_9", "Standard Identifiers", "F192", "ECU Hardware Number"),
    ("MCU_NM_ID_LIST_11", "Standard Identifiers", "F194", "ECU Software Number"),
    ("MCU_STD_DID_List_22", "Standard Identifiers", "0200", "Boot Flag"),
]

# Comprehensive Req. ID mapping without hardcoded statuses
REQ_ID_MAPPING = {
    "Standard Identifiers": {
        "Boot Software Identification": ("MCU_STD_DID_List_1", "F180"),
        "Application Software Identification": ("MCU_STD_DID_List_2", "F181"),
        "Application Data Identification": ("MCU_STD_DID_List_3", "F182"),
        "Vehicle Manufacturer spare part number": ("MCU_STD_DID_List_4", "F187"),
        "System Supplier Identifier": ("MCU_STD_DID_List_5", "F18A"),
        "ECU Manufacturing Date": ("MCU_STD_DID_List_6", "F18B"),
        "ECU Serial Number": ("MCU_STD_DID_List_7", "F18C"),
        "VIN-Vehicle Identification Number": ("MCU_STD_DID_List_8", "F190"),
        "System Supplier ECU Hardware Version Number": ("MCU_STD_DID_List_10", "F193"),
        "System Supplier ECU Software Version Number": ("MCU_STD_DID_List_12", "F195"),
        "System Name/Engine Type": ("MCU_STD_DID_List_13", "F197"),
        "Repair Shop Code/Tester Serial Number": ("MCU_STD_DID_List_14", "F198"),
        "Programming Date": ("MCU_STD_DID_List_15", "F199"),
        "ECU Installation Date": ("MCU_STD_DID_List_16", "F19D"),
        "System Supplier part number": ("MCU_STD_DID_List_17", "F1F0"),
        "Model number": ("MCU_STD_DID_List_18", "0100"),
        "Variant Code": ("MCU_STD_DID_List_19", "0101"),
        "Feature Code": ("MCU_STD_DID_List_20", "0102"),
        "Active Diagnostic Session": ("MCU_STD_DID_List_21", "F186"),
        "HISTORY ZONE": ("MCU_STD_DID_List_23", "0201"),
        "Engine State": ("MCU_NM_ID_LIST_4", "F1BE"),
        "ECU Hardware Number": ("MCU_NM_ID_LIST_9", "F192"),
        "ECU Software Number": ("MCU_NM_ID_LIST_11", "F194"),
        "Access Timing Parameters": ("Unknown", "0304"),
    },
    "Generic ECU Read": {
        "Odometer": ("MCU_GenECU_Read_list_1", "F1B0"),
        "Battery Voltage": ("MCU_GenECU_Read_list_2", "F1B1"),
        "Vehicle Speed": ("MCU_GenECU_Read_list_3", "F1B2"),
        "Motor speed": ("MCU_GenECU_Read_list_4", "F1B3"),
        "Reset Reason": ("MCU_GenECU_Read_list_6", "F1B4"),
        "Reset Counter": ("MCU_GenECU_Read_list_7", "F1B5"),
        "Ignition Counter": ("MCU_GenECU_Read_list_8", "F1B6"),
    },
    "Network Mgmnt": {
        "VCU_100 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU_100 Timeout Healing Time": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU3_100 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU3_100 Healing Time Threshold": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU5_500 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU5_500 Healing Time": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU8_10 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU8_10 Healing Time": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU9_10 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU9_10 Healing Time": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU14_20 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU14_20 Healing Time": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU17_500 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU17_500 Healing Time": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU5_100 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU_ACTIVE Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "VCU_ACTIVE Healing Time": ("MCU_NM_ID_LIST_1", "F1D2"),
        "BMS5_10 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "BMS5_10 Healing Time": ("MCU_NM_ID_LIST_1", "F1D2"),
        "BMS6_10 Timeout": ("MCU_NM_ID_LIST_1", "F1D2"),
        "BMS6_10 Healing Time": ("MCU_NM_ID_LIST_1", "F1D2"),
        "Network Management Enable": ("MCU_NM_ID_LIST_3", "F1D4"),
        "CAN Wakeup Msg Count": ("MCU_NM_ID_LIST_5", "0104"),
        "Quick Recovery time out period (in WAKEUP)": ("MCU_NM_ID_LIST_5", "0104"),
        "Quick Recovery retry limit (in WAKEUP)": ("MCU_NM_ID_LIST_5", "0104"),
        "Slow Recovery timeout period (in WAKEUP)": ("MCU_NM_ID_LIST_5", "0104"),
        "Slow Recovery retry limit (in WAKEUP)": ("MCU_NM_ID_LIST_5", "0104"),
    },
    "Network Mismatch": {
        "VCU_100 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU_100 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU_100 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU_100 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU3_100 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU3_100 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU3_100 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU3_100 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU3_100 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU3_100 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU5_500 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU5_500 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU5_500 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU5_500 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU8_10 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU8_10 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU8_10 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU8_10 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU9_10 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU9_10 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU9_10 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU9_10 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU9_10 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU9_10 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU14_20 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU14_20 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU14_20 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU14_20 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU17_500 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU17_500 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU17_500 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU17_500 DLC Mismatch Healing Threshold (Dematuration)": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU17_500 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU17_500 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU3_100 Alive Counter Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU3_100 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU9_10 Alive Counter Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU9_10 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU17_500 Alive Counter Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU17_500 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU_Active DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU_Active Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU_Active DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU_Active Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU5_500 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU5_500 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU5_500 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "VCU5_500 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS5_10 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS5_10 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS5_10 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS5_10 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS5_10 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS5_10 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS5_10 Alive Counter Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS5_10 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS6_10 DLC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS6_10 CRC Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS6_10 Parity Mismatch Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS6_10 DLC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS6_10 CRC Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS6_10 Parity Mismatch Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS6_10 Alive Counter Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
        "BMS6_10 Alive Counter Healing Threshold": ("MCU_NM_ID_LIST_2", "F1D3"),
    },
    "Can Configuration": {
        "Critical CAN Signal Invalid Time": ("MCU_NM_ID_LIST_4", "F1D5"),
        "Main CAN Bus Off Healing Time": ("MCU_NM_ID_LIST_4", "F1D5"),
        "CAN Timeout Since Powerup": ("MCU_NM_ID_LIST_4", "F1D5"),
        "CAN Wakeup Feature Enable": ("MCU_NM_ID_LIST_4", "0103"),
        "NM Drive Cnt to Clear DTC": ("MCU_NM_ID_LIST_4", "0103"),
        "Busoff Fast Recovery Time": ("MCU_NM_ID_LIST_4", "0103"),
        "Fast Bus off Recovery Count": ("MCU_NM_ID_LIST_4", "0103"),
        "Busoff Slow Recovery Time": ("MCU_NM_ID_LIST_4", "0103"),
        "NM IGN On Startup Delay": ("MCU_NM_ID_LIST_4", "0103"),
        "NM Restart Dly Time After Under Vol Recovery": ("MCU_NM_ID_LIST_4", "0103"),
        "NM Restart Dly Time After Over Vol Recovery": ("MCU_NM_ID_LIST_4", "0103"),
        "NM Restart Dly Time After Bus Off recovery": ("MCU_NM_ID_LIST_4", "0103"),
        "NM Restart Dly Time After Cranking": ("MCU_NM_ID_LIST_4", "0103"),
    },
    "Faults Configuration": {
        "Motor Over Temp Fault Detection": ("MCU_Fun_ID_List_1", "078F"),
        "Low Temperature Faults Healing Hysteresis": ("MCU_Fun_ID_List_1", "078F"),
        "Motor Low Temp Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Motor Low Temp Fault Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "MCU Over Temp Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "MCU Low Temp Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "MCU Low Temp Fault Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "Cooling Plate Over Temp Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Cooling Plate Temp Sensor Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "uC High Temp Fault Declaration Threshold": ("MCU_Fun_ID_List_1", "078F"),
        "uC Over Temp Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "MCU Temp Sensors Plausibility Failed Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Phases Overcurrent Peak Faults Absolute Declaration Threshold": ("MCU_Fun_ID_List_1", "078F"),
        "Phases Overcurrent Peak Faults Deviation Declaration Threshold": ("MCU_Fun_ID_List_1", "078F"),
        "Phase Overcurrent Peak Faults Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Phase Sensor Invalid Faults Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Phase Disconnected Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Active Short Circuit Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Active Short Circuit Fault Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "Gate Drivers Output Feedback Active CBIT Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Gate Drivers Output Feedback Active CBIT Fault Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "Gate Drivers Output Feedback Inactive CBIT Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Gate Drivers Output Feedback Inactive CBIT Fault Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "BAT OverCurrent Fault Absolute Declaration Threshold": ("MCU_Fun_ID_List_1", "078F"),
        "BAT OverCurrent Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "BAT Current Sensor Invalid Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "BAT Low Voltage Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "BAT High Voltage Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "BAT Under Voltage Fault Declaration Threshold": ("MCU_Fun_ID_List_1", "078F"),
        "BAT Under Voltage Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Battery_Overvoltage": ("MCU_Fun_ID_List_1", "078F"),
        "BAT Over Voltage Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "BAT Voltage Sensing Plausibility Failed Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Active Discharge Feedback CBIT Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Active Discharge Feedback CBIT Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "Active Discharge Timeout Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Active Discharge Timeout Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "Gate Drivers Under Voltage Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Gate Drivers Fault Detection": ("MCU_Fun_ID_List_1", "078F"),
        "KL30 Under Voltage": ("MCU_Fun_ID_List_1", "078F"),
        "KL30 Under Voltage Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "KL30 Over Voltage": ("MCU_Fun_ID_List_1", "078F"),
        "KL30 Over Voltage Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "KL30 LoS Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Motor Position Sensor Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Motor Speed Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Overspeed Threshold": ("MCU_Fun_ID_List_1", "078F"),
        "Speed_upper_limit": ("MCU_Fun_ID_List_1", "078F"),
        "Overspeed_limit_Margin": ("MCU_Fun_ID_List_1", "078F"),
        "Speed_lower_limit": ("MCU_Fun_ID_List_1", "078F"),
        "Overspeed_lower_Margin": ("MCU_Fun_ID_List_1", "078F"),
        "Motor Overspeed Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "Motor Stall Timeout": ("MCU_Fun_ID_List_1", "078F"),
        "IO Expander Configuration Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "IO Expander Configuration Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "CAN Transceiver Mode Detection Time": ("MCU_Fun_ID_List_1", "078F"),
        "CAN Transceiver Mode Healing Time": ("MCU_Fun_ID_List_1", "078F"),
        "Sensors 5V Fault Detection Time": ("MCU_Fun_ID_List_1", "078F"),
    },
    "True Drive Parameters": {
        "Angle Offset": ("MCU_Fun_ID_List_2", "0790"),
        "Angle Offset Delay": ("MCU_Fun_ID_List_2", "0790"),
        "Motor Temp - Min Cut": ("MCU_Fun_ID_List_2", "0790"),
        "Motor Temp - Min Start": ("MCU_Fun_ID_List_2", "0790"),
        "Motor Temp - Max Start": ("MCU_Fun_ID_List_2", "0790"),
        "Motor Temp - Max Cut": ("MCU_Fun_ID_List_2", "0790"),
        "Controller Temp - Min Cut": ("MCU_Fun_ID_List_2", "0790"),
        "Controller Temp - Min Start": ("MCU_Fun_ID_List_2", "0790"),
        "Controller Temp - Max Start": ("MCU_Fun_ID_List_2", "0790"),
        "Controller Temp - Max Cut": ("MCU_Fun_ID_List_2", "0790"),
        "Cooling Plate Temp - Min Cut": ("MCU_Fun_ID_List_2", "0790"),
        "Cooling Plate Temp - Min Start": ("MCU_Fun_ID_List_2", "0790"),
        "Cooling Plate Temp - Max Start": ("MCU_Fun_ID_List_2", "0790"),
        "Cooling Plate Temp - Max Cut": ("MCU_Fun_ID_List_2", "0790"),
        "Bus Voltage - Min Cut": ("MCU_Fun_ID_List_2", "0790"),
        "Bus Voltage - Min Start": ("MCU_Fun_ID_List_2", "0790"),
        "Bus Voltage - Max Start": ("MCU_Fun_ID_List_2", "0790"),
        "Bus Voltage - Max Cut": ("MCU_Fun_ID_List_2", "0790"),
        "Max Bus Current Limit Activation": ("MCU_Fun_ID_List_2", "0790"),
        "Max Bus Current Limit": ("MCU_Fun_ID_List_2", "0790"),
        "Min Bus Current Limit Activation": ("MCU_Fun_ID_List_2", "0790"),
        "Min Bus Current Limit": ("MCU_Fun_ID_List_2", "0790"),
        "Max Bus Voltage Limit Activation": ("MCU_Fun_ID_List_2", "0790"),
        "Max Bus Voltage Limit": ("MCU_Fun_ID_List_2", "0790"),
        "Min Bus Voltage Limit Activation": ("MCU_Fun_ID_List_2", "0790"),
        "Min Bus Voltage Limit": ("MCU_Fun_ID_List_2", "0790"),
        "Battery Maximum Current": ("MCU_Fun_ID_List_2", "0790"),
        "Bus Under-Voltage": ("MCU_Fun_ID_List_2", "0790"),
        "Bus Over-Voltage": ("MCU_Fun_ID_List_2", "0790"),
    },
    "ECU Identifiers": {
        "MCU EOL Calibration / Resolver Offset Calibration information": ("MCU_ECU_ID_list_1", "1500"),
    },
    "Freeze Frame": {
        "Freeze Frame Data": ("MCU_FF_ID_List_1", "078E"),
        "Snapshot Data": ("MCU_FF_ID_List_2", "F1B9"),
    },
    "Routine Control": {
        "Active Discharge": ("Unknown", "0296"),
        "Resolver Autocalibration": ("Unknown", "0295"),
    }
}

def validate_file_path(file_path, file_description):
    if not os.path.isfile(file_path):
        raise ValueError(f"{file_description} not found at: {file_path}")
    return file_path

def ensure_output_directory(output_file):
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created directory: {output_dir}")

def normalize_service_name(service):
    service = str(service).strip().upper() if service else ""
    return service

def strip_prefix(service):
    service = str(service).strip()
    match = re.match(r'^([0-9A-Fa-f]{3,4})\s+(.+)$', service)
    return (match.group(1).upper(), match.group(2).strip()) if match else (None, service)

def is_valid_did(did):
    did = str(did).strip().upper()
    return did == "0x31" or bool(re.match(r'^[0-9A-F]{3,4}$', did))

def find_column_index(headers, possible_names):
    headers = [str(h).lower() if h else "" for h in headers]
    for name in possible_names.split(";"):
        name = name.lower().strip()
        if name in headers:
            return headers.index(name) + 1
    return None

def extract_services_from_srd(file_path, extracted_srd_path, sheet_names=None):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        print(f"\nSRD file: {file_path}")
        print(f"Available sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"Error loading SRD file: {e}")
        return set(), {}, []

    services = set()
    services_with_original = {}
    services_with_details = []

    extracted_wb = openpyxl.Workbook()
    extracted_s = extracted_wb.active
    extracted_s.title = "Extracted SRD Data"
    headers = ["Sheet", "Group", "LID", "Description", "Req. ID", "Identifier"]
    for col_idx, header in enumerate(headers, 1):
        extracted_s.cell(row=1, column=col_idx).value = header
        extracted_s.cell(row=1, column=col_idx).font = Font(bold=True)

    extracted_row_idx = 2
    sheets_to_process = sheet_names if sheet_names else ["DID", "NM", "Functional Identifiers"]
    sheets_to_process = [s for s in sheets_to_process if s in workbook.sheetnames]

    for s_name in sheets_to_process:
        sheet = workbook[s_name]
        print(f"\nProcessing SRD sheet: '{s_name}'")
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        print(f"Header row: {header_row}")

        desc_col = find_column_index(header_row, "Description;Parameter")
        req_col = find_column_index(header_row, "Req;Req. ID;Requirement")
        group_col = find_column_index(header_row, "Group;Function;Group / Function")
        lid_col = find_column_index(header_row, "LID;DID")
        id_col = find_column_index(header_row, "Identifier;DID")
        print(f"Detected columns: Description={desc_col}, Req. ID={req_col}, Group={group_col}, LID={lid_col}, Identifier={id_col}")

        if not (desc_col and id_col):
            print(f"Skipping sheet '{s_name}': Missing required columns")
            continue

        row_count = 0
        for row in sheet.iter_rows(min_row=2, max_col=max(filter(None, [desc_col, req_col, group_col, lid_col, id_col])), values_only=True):
            row_count += 1
            group = str(row[group_col - 1] or "").strip() if group_col and len(row) >= group_col else ""
            lid = str(row[lid_col - 1] or "").strip() if lid_col and len(row) >= lid_col else ""
            service = str(row[desc_col - 1] or "").strip() if desc_col and len(row) >= desc_col else ""
            req_id = str(row[req_col - 1] or "").strip() if req_col and len(row) >= req_col else ""
            identifier = str(row[id_col - 1] or "").strip().upper() if id_col and len(row) >= id_col else ""

            print(f"SRD Row {row_count + 1} in {s_name}: Service={service}, Req. ID={req_id}, Identifier={identifier}, Group={group}, LID={lid}")
            if not (service and (identifier or lid)):
                print(f"SRD Row {row_count + 1} in {s_name}: Skipped (invalid service={service}, identifier={identifier}, lid={lid})")
                continue

            normalized_service = normalize_service_name(service)
            if normalized_service:
                services.add(normalized_service)
                services_with_original[normalized_service] = service
                services_with_details.append((group, lid, service, s_name, req_id, identifier))
                extracted_s.cell(row=extracted_row_idx, column=1).value = s_name
                extracted_s.cell(row=extracted_row_idx, column=2).value = group
                extracted_s.cell(row=extracted_row_idx, column=3).value = lid
                extracted_s.cell(row=extracted_row_idx, column=4).value = service
                extracted_s.cell(row=extracted_row_idx, column=5).value = req_id
                extracted_s.cell(row=extracted_row_idx, column=6).value = identifier
                extracted_row_idx += 1
            else:
                print(f"SRD Row {row_count + 1} in {s_name}: Skipped empty normalized service")

        print(f"Processed {row_count} row(s) in SRD sheet '{s_name}'")

    for col in extracted_s.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        extracted_s.column_dimensions[col[0].column_letter].width = max_length + 2

    ensure_output_directory(extracted_srd_path)
    try:
        extracted_wb.save(extracted_srd_path)
        print(f"Extracted SRD services saved to: {extracted_srd_path}")
    except Exception as e:
        print(f"Error during saving extracted SRD data: {e}")

    print(f"\nSRD details: {services_with_details}")
    return services, services_with_original, services_with_details

def extract_log_data(log_file_path, sheet_name=None):
    try:
        workbook = openpyxl.load_workbook(log_file_path, data_only=True)
        print(f"\nLog file: {log_file_path}")
        print(f"Available sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"Error loading log file: {e}")
        return {}, {}, {}, {}

    if not workbook.sheetnames:
        print(f"Error: No sheets found in {log_file_path}")
        return {}, {}, {}, {}

    if not sheet_name:
        sheet_name = workbook.sheetnames[0]
        print(f"No sheet name specified. Using first sheet: '{sheet_name}'")

    if sheet_name not in workbook.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
        return {}, {}, {}, {}

    sheet = workbook[sheet_name]
    log_data = {}
    log_original_names = {}
    log_dids = {}
    log_groups = {}

    print(f"\nProcessing log sheet: '{sheet_name}'")
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    print(f"Header row: {header_row}")

    group_col = find_column_index(header_row, "File Name;Group;Function;CID")
    service_col = find_column_index(header_row, "DID / Sub-service;Description;Service")
    status_col = find_column_index(header_row, "Status")
    print(f"Detected columns: Group={group_col}, Description={service_col}, Status={status_col}")

    if not (service_col and status_col):
        print(f"Skipping sheet '{sheet_name}': Missing required columns")
        return {}, {}, {}, {}

    row_count = 0
    for row in sheet.iter_rows(min_row=2, max_col=max(filter(None, [group_col, service_col, status_col])), values_only=True):
        row_count += 1
        group = str(row[group_col - 1] or "").strip() if group_col and len(row) >= group_col else ""
        service = str(row[service_col - 1] or "").strip() if service_col and len(row) >= service_col else ""
        status = str(row[status_col - 1] or "").strip() if status_col and len(row) >= status_col else ""
        did, service_name = strip_prefix(service) if service else (None, None)

        print(f"Log Row {row_count + 1}: Service={service_name or service}, DID={did}, Status={status}, Group={group}")
        if not (service_name or service):
            print(f"Log Row {row_count + 1}: Skipped (no service)")
            continue

        normalized_service = normalize_service_name(service_name or service)
        if normalized_service:
            log_data[normalized_service] = status
            log_original_names[normalized_service] = service_name or service
            log_dids[normalized_service] = did or ""
            log_groups[normalized_service] = group
        else:
            print(f"Log Row {row_count + 1}: Skipped empty normalized service")

    print(f"Processed {row_count} row(s) in log sheet '{sheet_name}'")
    return log_data, log_original_names, log_dids, log_groups

def compare_and_generate_report(srd_services, srd_original_names, srd_details, log_data, log_original_names, log_dids, log_groups, output_file):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Implementation Report"

        headers = ["Req. ID", "Group / Function", "LID/DID", "Description", "Status"]
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx).value = header
            sheet.cell(row=1, column=col_idx).font = Font(bold=True)

        pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        not_impl_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        row_idx = 2
        processed_services = set()
        for req_id, group_name, did, service_name in EXPECTED_ORDER:
            normalized_service = normalize_service_name(service_name)
            processed_services.add(normalized_service)

            status = log_data.get(normalized_service, "Not Tested")
            if did.upper() in NON_IMPLEMENTED_DIDS or did.upper() == "0104":
                status = "Not Tested"

            sheet.cell(row=row_idx, column=1).value = req_id
            sheet.cell(row=row_idx, column=2).value = group_name
            sheet.cell(row=row_idx, column=3).value = did
            sheet.cell(row=row_idx, column=4).value = service_name
            sheet.cell(row=row_idx, column=5).value = status
            sheet.cell(row=row_idx, column=5).fill = pass_fill if status.lower() == "pass" else fail_fill if status.lower() == "failed" else not_impl_fill

            print(f"Output row {row_idx}: Req. ID={req_id}, Group={group_name}, LID/DID={did}, Service={service_name}, Status={status}")
            row_idx += 1

        for group_name, services in sorted(REQ_ID_MAPPING.items(), key=lambda x: x[0]):
            for service_name, (req_id, did) in sorted(services.items(), key=lambda x: x[1][0]):
                normalized_service = normalize_service_name(service_name)
                if normalized_service in processed_services:
                    print(f"Skipped REQ_ID_MAPPING: Service={service_name}, DID={did}, (already processed)")
                    continue

                status = log_data.get(normalized_service, "Not Tested")
                if did.upper() in NON_IMPLEMENTED_DIDS or did.upper() == "0104":
                    status = "Not Tested"

                sheet.cell(row=row_idx, column=1).value = req_id
                sheet.cell(row=row_idx, column=2).value = group_name
                sheet.cell(row=row_idx, column=3).value = did
                sheet.cell(row=row_idx, column=4).value = service_name
                sheet.cell(row=row_idx, column=5).value = status
                sheet.cell(row=row_idx, column=5).fill = pass_fill if status.lower() == "pass" else fail_fill if status.lower() == "failed" else not_impl_fill

                print(f"Output row {row_idx}: Req. ID={req_id}, Group={group_name}, LID/DID={did}, Service={service_name}, Status={status}")
                processed_services.add(normalized_service)
                row_idx += 1

        srd_by_did = {str(identifier).upper(): (group, service, req_id, lid) for group, lid, service, _, req_id, identifier in srd_details if identifier}
        unmatched_services = []
        for group, lid, service_name, s_name, req_id, identifier in sorted(srd_details, key=lambda x: (str(x[4]) or "", str(x[0]) or "", str(x[2]) or "")):
            normalized_service = normalize_service_name(service_name)
            identifier = str(identifier).upper() if identifier else ""
            if normalized_service in processed_services:
                print(f"Unmatched SRD in {s_name}: Skipped Service={service_name}, DID={identifier or lid} (already processed)")
                continue
            if not is_valid_did(identifier or lid):
                print(f"Unmatched SRD in {s_name}: Skipped Service={service_name}, DID={identifier or lid} (invalid DID)")
                continue

            status = log_data.get(normalized_service, "Not Tested")
            if identifier in NON_IMPLEMENTED_DIDS or identifier == "0104":
                status = "Not Tested"

            unmatched_services.append((req_id or "", group or "Unknown", identifier or lid, service_name, status))
            print(f"Unmatched SRD in {s_name}: Service={service_name}, DID={identifier or lid}, Status={status}")

        for req_id, group, did, service_name, status in sorted(unmatched_services, key=lambda x: x[2]):
            sheet.cell(row=row_idx, column=1).value = req_id
            sheet.cell(row=row_idx, column=2).value = group
            sheet.cell(row=row_idx, column=3).value = did
            sheet.cell(row=row_idx, column=4).value = service_name
            sheet.cell(row=row_idx, column=5).value = status
            sheet.cell(row=row_idx, column=5).fill = pass_fill if status.lower() == "pass" else fail_fill if status.lower() == "failed" else not_impl_fill
            print(f"SRD Row {row_idx}: Req. ID={req_id}, Group={group}, LID/DID={did}, Service={service_name}, Status={status}")
            row_idx += 1

        for col_idx in range(1, len(headers) + 1):
            max_length = max(len(str(sheet.cell(row=r, column=col_idx).value or "")) for r in range(1, row_idx))
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2

        ensure_output_directory(output_file)
        workbook.save(output_file)
        print(f"\nReport saved to: {output_file}")

    except Exception as e:
        print(f"Error in report generation: {e}")
        raise

def main():
    parser = argparse.ArgumentParser(description="Generate UDS compliance report")
    parser.add_argument("--srd-file", default=SRD_path, help="SRD Excel file path")
    parser.add_argument("--log-file", default=None, help="Log Excel file path")
    parser.add_argument("--output-file", default=OUTPUT_path, help="Output Excel report")
    parser.add_argument("--extracted-srd-file", default=EXTRACTED_SRD_path, help="Extracted SRD data path")
    parser.add_argument("--srd-sheets", default="DID,NM,Functional Identifiers", help="Comma-separated SRD sheet names")
    parser.add_argument("--log-sheet", default=None, help="Log sheet name")
    args = parser.parse_args()

    try:
        # Run output_with_raw.py unless --log-file is provided
        log_file = args.log_file
        if not log_file:
            print("\nRunning output_with_raw.py to generate log report...")
            try:
                log_file = output_with_raw.main()
                if log_file and not os.path.isfile(log_file):
                    print(f"Warning: output_with_raw.main() returned {log_file}, but file does not exist")
                    log_file = None
            except Exception as e:
                print(f"Failed to run output_with_raw.main(): {e}")
                log_file = None

            if not log_file:
                print("Falling back to subprocess execution of output_with_raw.py...")
                try:
                    result = subprocess.run(
                        [sys.executable, "output_with_raw.py"],
                        capture_output=True,
                        text=True,
                        check=True
                    )
                    print("output_with_raw.py output:", result.stdout)
                    if result.stderr:
                        print("output_with_raw.py errors:", result.stderr)
                except subprocess.CalledProcessError as e:
                    print(f"Subprocess failed: {e}")
                    print("Subprocess output:", e.output)
                    print("Subprocess errors:", e.stderr)

            # Check default path
            log_file = log_file or UDS_path
            if not os.path.isfile(log_file):
                raise FileNotFoundError(
                    f"Log file not found at {log_file}. Ensure output_with_raw.py generates the file or provide --log-file."
                )

        srd_file = validate_file_path(args.srd_file, "SRD file")
        log_file = validate_file_path(log_file, "Log file")

        print("\nExtracting SRD services...")
        srd_services, srd_original, srd_details = extract_services_from_srd(srd_file, args.extracted_srd_file, args.srd_sheets.split(","))

        print("\nExtracting log data...")
        log_data, log_original_names, log_dids, log_groups = extract_log_data(log_file, args.log_sheet)

        print("\nGenerating report...")
        compare_and_generate_report(srd_services, srd_original, srd_details, log_data, log_original_names, log_dids, log_groups, args.output_file)

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
