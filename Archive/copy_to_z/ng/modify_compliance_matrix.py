import argparse
import subprocess

import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import sys
import re
import output_with_raw
import ng
# Define file paths
# SRD_path = r"C:\Users\ilyar\Downloads\HD-UP-ICD-242601-UDID.xlsx"
# UDS_path = r"C:\Users\ilyar\PycharmProjects\UDS\Logs\03.01.11\03.01.11_report.xlsx"
# OUTPUT_path = os.path.join(os.path.dirname(UDS_path), "UDS_Compliance_matrix_UPP_v3.01.11.xlsx")
# EXTRACTED_SRD_path = os.path.join(os.path.dirname(UDS_path), "extracted_srd_data.xlsx")


username = os.environ.get('USERNAME', 'unknown')
if username == 'unknown':
    raise EnvironmentError("USERNAME environment variable not set.")

#Logs_folder = os.path.join("Logs")
###############################
base_log_dir = os.path.dirname(os.path.abspath(__file__))
Logs_folder = os.path.join(base_log_dir, "Logs")
os.makedirs(Logs_folder, exist_ok=True)
#####################################

print(base_log_dir)

result_folder = os.environ.get("RESULT_FOLDER")

if not result_folder:
    # Fallback: find the most recently created folder inside Logs
    logs_base = os.path.join(base_log_dir, "Logs")
    subfolders = [
        os.path.join(logs_base, d) for d in os.listdir(logs_base)
        if os.path.isdir(os.path.join(logs_base, d))
    ]

    if not subfolders:
        raise ValueError("No subfolders found in Logs to infer RESULT_FOLDER")

    latest_folder = max(subfolders, key=os.path.getctime)
    result_folder = os.path.basename(latest_folder)
    print(f"Fallback: Using latest created result folder → {result_folder}")
####    raise ValueError("RESULT_FOLDER environment variable not set")

UDS_path = os.path.join(base_log_dir,"Logs", result_folder,  f"{result_folder}_report.xlsx")
OUTPUT_path = os.path.join(base_log_dir,"Logs", result_folder, f"UDS_Compliance_matrix_NewGen_v{result_folder}.xlsx")
EXTRACTED_SRD_path = os.path.join(base_log_dir,"Logs", result_folder,  "extracted_srd_data.xlsx")
SRD_path = os.path.join(base_log_dir, "Documents", "New Gen D-6 Microcontroller UDS DIDs.xlsx")


# Define non-implemented DIDs
NON_IMPLEMENTED_DIDS = {"F1BE"}

# Expected output order for first 10 rows
EXPECTED_ORDER = [
#     ("MCU_STD_DID_List_1", "Standard Identifiers", "F180", "Boot Software Identification"),
#     ("MCU_GenECU_Read_list_1", "Generic ECU Read", "F1B0", "Odometer"),
#     ("MCU_NM_ID_LIST_1", "Network Mgmnt", "F1D2", "VCU3_100 Timeout"),
#     ("MCU_NM_ID_LIST_4", "Can Configuration", "F1D5", "Critical CAN Signal Invalid Time"),
#     ("MCU_Fun_ID_List_1", "Faults Configuration", "078F", "Motor Over Temp Fault Detection"),
# #    ("Unknown", "Routine Control", "0296", "Active Discharge"),
#     ("MCU_NM_ID_LIST_4", "Standard Identifiers", "F1BE", "Engine State"),
     ("MCU_NM_ID_LIST_9", "Standard Identifiers", "B11B", "Id_Act"),
     ("MCU_NM_ID_LIST_11", "Standard Identifiers", "B11C", "Iq_Act"),
      ("MCU_STD_DID_List_22", "Standard Identifiers", "B11D", "VLL_rms"),

]
# Comprehensive Req. ID mapping without hardcoded statuses
REQ_ID_MAPPING = {
    "Standard Identifiers": {
            "BootSoftwareIdentificationDataIdentifier": ("MCU_STD_DID_List_1", "F180"),
            "ApplicationSoftwareIdentificationDataIdentifier": ("MCU_STD_DID_List_2", "F181"),
            "ApplicationDataIdentification": ("MCU_STD_DID_List_3", "F182"),
            "ActiveDiagnosticSessionDataIdentifier": ("MCU_STD_DID_List_4", "F186"),
            "SystemSupplierIdentifierDataIdentifier": ("MCU_STD_DID_List_5", "F18A"),
            "ECUManufacturingDateDataIdentifier": ("MCU_STD_DID_List_6", "F18B"),
            "ECUSerialNumberDataIdentifier": ("MCU_STD_DID_List_7", "F18C"),
            "VINDataIdentifier": ("MCU_STD_DID_List_8", "F190"),
            "vehicleManufacturerECUHardwareNumberDataIdentifier": ("MCU_STD_DID_List_9", "F191"),
            "systemSupplierECUHardwareNumberDataIdentifier": ("MCU_STD_DID_List_10", "F192"),
            "systemSupplierECUHardwareVersionNumberDataIdentifier": ("MCU_STD_DID_List_11", "F193"),
            "systemSupplierECUSoftwareNumberDataIdentifier": ("MCU_STD_DID_List_12", "F194"),
            "systemSupplierECUSoftwareVersionNumberDataIdentifier": ("MCU_STD_DID_List_13", "F195"),
            "systemNameOrEngineTypeDataIdentifier": ("MCU_STD_DID_List_14", "F197"),
            "repairShopCodeOrTesterSerialNumberDataIdentifier": ("MCU_STD_DID_List_15", "F198"),
            "programmingDateDataIdentifier": ("MCU_STD_DID_List_16", "F199"),
            "calibrationDateDataIdentifier": ("MCU_STD_DID_List_17", "F19B"),
            "calibrationEquipmentSoftwareNumberDataIdentifier": ("MCU_STD_DID_List_18", "F19C"),
            "ECUInstallationDateDataIdentifier": ("MCU_STD_DID_List_19", "F19D"),
            "ODXFileDataIdentifier": ("MCU_STD_DID_List_20", "F19E"),
            "LMM Specific": ("MCU_STD_DID_List_21", "F101"),
            "Boot Flag": ("MCU_STD_DID_List_22", "0200"),
            "History Zone": ("MCU_STD_DID_List_23", "0201"),
            "Access Timing Parameters": ("MCU_Special_DID_List_1", "0304")
        },


    "Network Mgmnt": {
        "Node_Absent_Detection_Threshold": ("MCU_NM_ID_LIST_1", "C000"),
        "Node_Absent_Recovery_Threshold": ("MCU_NM_ID_LIST_2", "C000"),
        "Message_Timeout_Detection_Threshold": ("MCU_NM_ID_LIST_3", "C000"),
        "Message_Timeout_Recovery_Threshold": ("MCU_NM_ID_LIST_4", "C000"),
        "NM_Drive_Cnt_to_Clear_DTC": ("MCU_NM_ID_LIST_5", "C000"),
        "Busoff_Fast_Recovery_Time": ("MCU_NM_ID_LIST_6", "C000"),
        "Fast_Bus_off_Recovery_Count": ("MCU_NM_ID_LIST_7", "C000"),
        "Busoff_Slow_Recovery_Time": ("MCU_NM_ID_LIST_8", "C000"),
        "NM_IGN_On_Startup_Delay": ("MCU_NM_ID_LIST_9", "C000"),
        "NM_Restart_Dly_Time_After_Under_Vol_Recovery": ("MCU_NM_ID_LIST_10", "C000"),
        "NM_Restart_Dly_Time_After_Over_Vol_Recovery": ("MCU_NM_ID_LIST_11", "C000"),
        "NM_Restart_Dly_Time_After_Bus_Off_recovery": ("MCU_NM_ID_LIST_12", "C000"),
        "CAN_Wakeup_Feature_Enable": ("MCU_NM_ID_LIST_13", "C000")
        # F1D2 - Message Timeout and Heal
    },


    "True Drive Parameters": {
        # All use MCU_Fun_ID_List_2
        "Motor Temp - Max Start": ("MCU_Special_DID_List_2", "B100"),
        "Motor Temp - Max Cut": ("MCU_Special_DID_List_3", "B101"),
        "Bus Voltage - Max Start": ("MCU_Special_DID_List_4", "B102"),
        "Bus Voltage - Max Cut": ("MCU_Special_DID_List_5", "B103"),
        "Bus Voltage - Min Start": ("MCU_Special_DID_List_6", "B104"),
        "Bus Voltage - Min Cut": ("MCU_Special_DID_List_7", "B105"),
        "Hill Assisst Activation": ("MCU_Special_DID_List_8", "B106"),
        "Assist Hill Hold Timeout": ("MCU_Special_DID_List_9", "B107"),
        "Assist Rollback Speed": ("MCU_Special_DID_List_10", "B108"),
        "Assist Brake Hold Time": ("MCU_Special_DID_List_11", "B109"),
        "Assist Minimum Torque": ("MCU_Special_DID_List_12", "B10A"),
        "Assist Brake Press Debounce Time": ("MCU_Special_DID_List_13", "B10B"),
        "Assist Drive Timeout Time": ("MCU_Special_DID_List_14", "B10C"),
        "Angle Offset": ("MCU_Special_DID_List_15", "B10D"),
        "Forward Max Speed Limit Value": ("MCU_Special_DID_List_16", "B10E"),
        "Reverse Max Speed Limit Value": ("MCU_Special_DID_List_17", "B10F"),
        "Actual Motor Speed": ("MCU_Special_DID_List_18", "B110"),
        "Actual Torque": ("MCU_Special_DID_List_19", "B111"),
        "MCU_Counter1_20": ("MCU_Special_DID_List_20", "B112"),
        "Motor_Current_Measured": ("MCU_Special_DID_List_21", "B113"),
        "MCU_Board_temp": ("MCU_Special_DID_List_22", "B114"),
        "MCU_IP_DCBus_Voltage": ("MCU_Special_DID_List_23", "B116"),
        "MCU_Operating_Mode": ("MCU_Special_DID_List_24", "B117"),
        "Motor_Power_Measured": ("MCU_Special_DID_List_25", "B118"),
        "Motor_Winding-Temp": ("MCU_Special_DID_List_26", "B119"),
        "MCU_IP_DC_Current": ("MCU_Special_DID_List_27", "B11A"),
        "Id_Act": ("MCU_Special_DID_List_28", "B11B"),
        "Iq_Act": ("MCU_Special_DID_List_29", "B11C"),
        "VLL_rms": ("MCU_Special_DID_List_30", "B11D"),
        "Motor Over Speed Threshold": ("MCU_Special_DID_List_31", "B11E"),
        "Motor Max Current Command": ("MCU_Special_DID_List_32", "B11F"),
        "Controller Temp - Max Start": ("MCU_Special_DID_List_33", "B120"),
        "Controller Temp - Max Cut": ("MCU_Special_DID_List_34", "B121"),
        "Controller Temp - Min Start": ("MCU_Special_DID_List_35", "B122"),
        "Controller Temp - Min Cut": ("MCU_Special_DID_List_36", "B123"),
        "Motor Temp - Min Start": ("MCU_Special_DID_List_37", "B124"),
        "Motor Temp - Min Cut": ("MCU_Special_DID_List_38", "B125"),
        "Max Bus Drive Current Limit Value": ("MCU_Special_DID_List_39", "B126"),
        "Max Bus Regen Current Limit Value": ("MCU_Special_DID_List_40", "B127"),
        "ADC Current Factor, Phase U": ("MCU_Special_DID_List_41", "B128"),
        "ADC Current Factor, Phase V": ("MCU_Special_DID_List_42", "B129"),
        "ADC Current Factor, Bus": ("MCU_Special_DID_List_43", "B12A"),
        "Ramp Up Drive": ("MCU_Special_DID_List_44", "B12B"),
        "Ramp Up Regen": ("MCU_Special_DID_List_45", "B12C"),
        "Ramp Down": ("MCU_Special_DID_List_46", "B12D"),
        "Power Max": ("MCU_Special_DID_List_47", "B12E"),
        "Rotor Position": ("MCU_Special_DID_List_48", "B12F"),
        "Frequency": ("MCU_Special_DID_List_49", "B130"),
        "Network Mgmt Configuration": ("MCU_Network_Mgmt_List_1", "C000")

    },
    "Freeze Frame": {
        "Freeze_Frame_Global_Identifiers": ("MCU_Freeze_Frame_List_1", "A8FF"),
        "Freeze_Frame_MCU": ("MCU_Freeze_Frame_List_2", "D5FF"),
    },
    "Routine Control": {
        # "Start Active Discharge": ("Unknown", "0296"),
        # "Start Resolver Autocalibration": ("Unknown", "0295"),
        # "Start History Zone Update": ("Unknown", "0201"),
        # "Start Compare CS": ("Unknown", "FF01"),
        # "Stop Active Discharge": ("Unknown", "0296"),
        # "Stop Resolver Autocalibration": ("Unknown", "0295"),
        # "Result Active Discharge": ("Unknown", "0296"),
        # "Result Resolver Autocalibration": ("Unknown", "0295"),
        # "Result History Zone Update": ("Unknown", "0201"),
        # "Result Compare CS": ("Unknown", "FF01"),
    }
}

def validate_file_path(file_path, file_description):
    if not os.path.isfile(file_path):
        raise ValueError(f"{file_description} not found at: {file_path}")
    return file_path

def ensure_output_directory(output_file):
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created directory: {output_dir}")

def normalize_service_name(service):
    if not service:
        return ""
    service = str(service).strip()
    # Remove common prefixes and suffixes
    prefixes = [
        r'^MISMATCH TX AND RX\s+',  # Remove "Mismatch Tx and Rx"
        r'^\s*[0-9A-F]{3,4}\s+'    # Remove leading DID
    ]
    suffixes = [
        r'\s+WRONG OUTPUT FAIL$',   # Remove "wrong output Fail"
        r'\s+FAIL$'                 # Remove trailing "Fail"
    ]
    for prefix in prefixes:
        service = re.sub(prefix, '', service, flags=re.IGNORECASE)
    for suffix in suffixes:
        service = re.sub(suffix, '', service, flags=re.IGNORECASE)
    service = service.strip().upper()
    # Map known service name variations to standard names
    service_mapping = {
        "NORMAL MIN TIMEOUT TIMER": "NORMAL MIN TIMEOUT TIMER",
        "SLEEP WAIT TIMER": "SLEEP WAIT TIMER"
    }
    return service_mapping.get(service, service)

def strip_prefix(service):
    service = str(service).strip()
    # Match DID followed by service name, considering possible prefixes
    match = re.match(r'^(?:MISMATCH TX AND RX\s+)?([0-9A-Fa-f]{3,4})\s+(.+)$', service, re.IGNORECASE)
    if match:
        return (match.group(1).upper(), match.group(2).strip())
    return (None, service)

def is_valid_did(did):
    did = str(did).strip().upper()
    return did == "0x31" or bool(re.match(r'^[0-9A-F]{3,4}$', did))

def find_column_index(headers, possible_names):
    headers = [str(h).lower() if h else "" for h in headers]
    for name in possible_names.split(";"):
        name = name.lower().strip()
        if name in headers:
            return headers.index(name) + 1
    return None

def extract_services_from_srd(file_path, extracted_srd_path, sheet_names=None):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        print(f"\nSRD file: {file_path}")
        print(f"Available sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"Error loading SRD file: {e}")
        return set(), {}, []

    services = set()
    services_with_original = {}
    services_with_details = []

    extracted_wb = openpyxl.Workbook()
    extracted_s = extracted_wb.active
    extracted_s.title = "Extracted SRD Data"
    headers = ["Sheet", "Group", "LID", "Description", "Req. ID", "Identifier"]
    for col_idx, header in enumerate(headers, 1):
        extracted_s.cell(row=1, column=col_idx).value = header
        extracted_s.cell(row=1, column=col_idx).font = Font(bold=True)

    extracted_row_idx = 2
    sheets_to_process = sheet_names if sheet_names else ["DID", "NM", "Functional Identifiers"]
    sheets_to_process = [s for s in sheets_to_process if s in workbook.sheetnames]

    for s_name in sheets_to_process:
        sheet = workbook[s_name]
        print(f"\nProcessing SRD sheet: '{s_name}'")
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        print(f"Header row: {header_row}")

        desc_col = find_column_index(header_row, "Description;Parameter")
        req_col = find_column_index(header_row, "Req;Req. ID;Requirement")
        group_col = find_column_index(header_row, "Group;Function;Group / Function")
        lid_col = find_column_index(header_row, "LID;DID")
        id_col = find_column_index(header_row, "Identifier;DID")
        print(f"Detected columns: Description={desc_col}, Req. ID={req_col}, Group={group_col}, LID={lid_col}, Identifier={id_col}")

        if not (desc_col and id_col):
            print(f"Skipping sheet '{s_name}': Missing required columns")
            continue

        row_count = 0
        for row in sheet.iter_rows(min_row=2, max_col=max(filter(None, [desc_col, req_col, group_col, lid_col, id_col])), values_only=True):
            row_count += 1
            group = str(row[group_col - 1] or "").strip() if group_col and len(row) >= group_col else ""
            lid = str(row[lid_col - 1] or "").strip() if lid_col and len(row) >= lid_col else ""
            service = str(row[desc_col - 1] or "").strip() if desc_col and len(row) >= desc_col else ""
            req_id = str(row[req_col - 1] or "").strip() if req_col and len(row) >= req_col else ""
            identifier = str(row[id_col - 1] or "").strip().upper() if id_col and len(row) >= id_col else ""

            print(f"SRD Row {row_count + 1} in {s_name}: Service={service}, Req. ID={req_id}, Identifier={identifier}, Group={group}, LID={lid}")
            if not (service and (identifier or lid)):
                print(f"SRD Row {row_count + 1} in {s_name}: Skipped (invalid service={service}, identifier={identifier}, lid={lid})")
                continue

            normalized_service = normalize_service_name(service)
            if normalized_service:
                services.add(normalized_service)
                services_with_original[normalized_service] = service
                services_with_details.append((group, lid, service, s_name, req_id, identifier))
                extracted_s.cell(row=extracted_row_idx, column=1).value = s_name
                extracted_s.cell(row=extracted_row_idx, column=2).value = group
                extracted_s.cell(row=extracted_row_idx, column=3).value = lid
                extracted_s.cell(row=extracted_row_idx, column=4).value = service
                extracted_s.cell(row=extracted_row_idx, column=5).value = req_id
                extracted_s.cell(row=extracted_row_idx, column=6).value = identifier
                extracted_row_idx += 1
            else:
                print(f"SRD Row {row_count + 1} in {s_name}: Skipped empty normalized service")

        print(f"Processed {row_count} row(s) in SRD sheet '{s_name}'")

    for col in extracted_s.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        extracted_s.column_dimensions[col[0].column_letter].width = max_length + 2

    ensure_output_directory(extracted_srd_path)
    try:
        extracted_wb.save(extracted_srd_path)
        print(f"Extracted SRD services saved to: {extracted_srd_path}")
    except Exception as e:
        print(f"Error during saving extracted SRD data: {e}")

    print(f"\nSRD details: {services_with_details}")
    return services, services_with_original, services_with_details

def extract_log_data(log_file_path, sheet_name=None):
    try:
        workbook = openpyxl.load_workbook(log_file_path, data_only=True)
        print(f"\nLog file: {log_file_path}")
        print(f"Available sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"Error loading log file: {e}")
        return {}, {}, {}, {}

    if not workbook.sheetnames:
        print(f"Error: No sheets found in {log_file_path}")
        return {}, {}, {}, {}

    if not sheet_name:
        sheet_name = workbook.sheetnames[0]
        print(f"No sheet name specified. Using first sheet: '{sheet_name}'")

    if sheet_name not in workbook.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
        return {}, {}, {}, {}

    sheet = workbook[sheet_name]
    log_data = {}
    log_original_names = {}
    log_dids = {}
    log_groups = {}

    print(f"\nProcessing log sheet: '{sheet_name}'")
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    print(f"Header row: {header_row}")

    group_col = find_column_index(header_row, "File Name;Group;Function;CID")
    service_col = find_column_index(header_row, "DID / Sub-service;Description;Service")
    status_col = find_column_index(header_row, "Status")
    print(f"Detected columns: Group={group_col}, Description={service_col}, Status={status_col}")

    if not (service_col and status_col):
        print(f"Skipping sheet '{sheet_name}': Missing required columns")
        return {}, {}, {}, {}

    row_count = 0
    for row in sheet.iter_rows(min_row=2, max_col=max(filter(None, [group_col, service_col, status_col])), values_only=True):
        row_count += 1
        group = str(row[group_col - 1] or "").strip() if group_col and len(row) >= group_col else ""
        service = str(row[service_col - 1] or "").strip() if service_col and len(row) >= service_col else ""
        status = str(row[status_col - 1] or "").strip() if status_col and len(row) >= status_col else ""
        did, service_name = strip_prefix(service) if service else (None, None)

        print(f"Log Row {row_count + 1}: Service={service_name or service}, DID={did}, Status={status}, Group={group}")
        if not (service_name or service):
            print(f"Log Row {row_count + 1}: Skipped (no service)")
            continue

        normalized_service = normalize_service_name(service_name or service)
        if normalized_service:
            log_data[normalized_service] = status
            log_original_names[normalized_service] = service_name or service
            log_dids[normalized_service] = did or ""
            log_groups[normalized_service] = normalize_group_name(group)
        else:
            print(f"Log Row {row_count + 1}: Skipped empty normalized service")

    print(f"Processed {row_count} row(s) in log sheet '{sheet_name}'")
    return log_data, log_original_names, log_dids, log_groups

def normalize_group_name(group):
    if not group:
        return None
    group = str(group).strip()
    # Map inconsistent group names from log report to standard names
    group_mapping = {
        "Generoid_ECU_F": "Generic ECU Read",
        "Network_Mismatch_F1D3": "Network Mismatch",
        "Network_Timeout_F1D2": "Network Mgmnt",
        "Network_F1D5": "Can Configuration",
        "Network_103": "Can Configuration",
        "TrueDrive_M": "True Drive Parameters",
        "Faults_C": "Faults Configuration",
        "Standard_I": "Standard Identifiers",
        "Routine_C": "Routine Control",
    }
    return group_mapping.get(group, group)

def compare_and_generate_report(srd_services, srd_original_names, srd_details, log_data, log_original_names, log_dids, log_groups, output_file):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Implementation Report"

        headers = ["Req. ID", "Group / Function", "LID/DID", "Description", "Status"]
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx).value = header
            sheet.cell(row=1, column=col_idx).font = Font(bold=True)

        pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        not_impl_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        row_idx = 2
        # Process EXPECTED_ORDER
        processed_services = set()
        for req_id, group_name, did, service_name in EXPECTED_ORDER:
            normalized_service = normalize_service_name(service_name)
            processed_services.add(normalized_service)

            status = log_data.get(normalized_service, "Not Tested")
            if did.upper() in NON_IMPLEMENTED_DIDS or did.upper() == "0104":
                status = "Not Tested"
            # Map 'Fail' to 'Failed' for output
            display_status = "Failed" if status.lower() == "fail" else status

            sheet.cell(row=row_idx, column=1).value = req_id
            sheet.cell(row=row_idx, column=2).value = group_name
            sheet.cell(row=row_idx, column=3).value = did
            sheet.cell(row=row_idx, column=4).value = service_name
            sheet.cell(row=row_idx, column=5).value = display_status
            sheet.cell(row=row_idx, column=5).fill = pass_fill if status.lower() == "pass" else fail_fill if status.lower() == "fail" else not_impl_fill

            print(f"Output row {row_idx}: Req. ID={req_id}, Group={group_name}, LID/DID={did}, Service={service_name}, Status={display_status}")
            row_idx += 1

        # Process REQ_ID_MAPPING entries
        for group_name, services in sorted(REQ_ID_MAPPING.items(), key=lambda x: x[0]):
            for service_name, (req_id, did) in sorted(services.items(), key=lambda x: x[1][0]):
                normalized_service = normalize_service_name(service_name)
                if normalized_service in processed_services:
                    print(f"Skipped REQ_ID_MAPPING: Service={service_name}, DID={did}, (already processed)")
                    continue

                status = log_data.get(normalized_service, "Not Tested")
                if did.upper() in NON_IMPLEMENTED_DIDS or did.upper() == "0104":
                    status = "Not Tested"
                # Map 'Fail' to 'Failed' for output
                display_status = "Failed" if status.lower() == "fail" else status

                sheet.cell(row=row_idx, column=1).value = req_id
                sheet.cell(row=row_idx, column=2).value = group_name
                sheet.cell(row=row_idx, column=3).value = did
                sheet.cell(row=row_idx, column=4).value = service_name
                sheet.cell(row=row_idx, column=5).value = display_status
                sheet.cell(row=row_idx, column=5).fill = pass_fill if status.lower() == "pass" else fail_fill if status.lower() == "fail" else not_impl_fill

                print(f"Output row {row_idx}: Req. ID={req_id}, Group={group_name}, LID/DID={did}, Service={service_name}, Status={display_status}")
                processed_services.add(normalized_service)
                row_idx += 1

        # Process unmatched SRD services
        srd_by_did = {str(identifier).upper(): (group, service, req_id, lid) for group, lid, service, _, req_id, identifier in srd_details if identifier}
        unmatched_services = []
        for group, lid, service_name, s_name, req_id, identifier in sorted(srd_details, key=lambda x: (str(x[4]) or "", str(x[0]) or "", str(x[2]) or "")):
            normalized_service = normalize_service_name(service_name)
            identifier = str(identifier).upper() if identifier else ""
            if normalized_service in processed_services:
                print(f"Unmatched SRD in {s_name}: Skipped Service={service_name}, DID={identifier or lid} (already processed)")
                continue
            if not is_valid_did(identifier or lid):
                print(f"Unmatched SRD in {s_name}: Skipped Service={service_name}, DID={identifier or lid} (invalid DID)")
                continue

            status = log_data.get(normalized_service, "Not Tested")
            if identifier in NON_IMPLEMENTED_DIDS or identifier == "0104":
                status = "Not Tested"
            # Map 'Fail' to 'Failed' for output
            display_status = "Failed" if status.lower() == "fail" else status

            unmatched_services.append((req_id or "", group or "Unknown", identifier or lid, service_name, display_status))
            print(f"Unmatched SRD in {s_name}: Service={service_name}, DID={identifier or lid}, Status={display_status}")

        # Add unmatched SRD services
        for req_id, group, did, service_name, display_status in sorted(unmatched_services, key=lambda x: x[2]):
            status = "Fail" if display_status.lower() == "failed" else display_status  # Convert back for fill logic
            sheet.cell(row=row_idx, column=1).value = req_id
            sheet.cell(row=row_idx, column=2).value = group
            sheet.cell(row=row_idx, column=3).value = did
            sheet.cell(row=row_idx, column=4).value = service_name
            sheet.cell(row=row_idx, column=5).value = display_status
            sheet.cell(row=row_idx, column=5).fill = pass_fill if status.lower() == "pass" else fail_fill if status.lower() == "fail" else not_impl_fill
            print(f"SRD Row {row_idx}: Req. ID={req_id}, Group={group}, LID/DID={did}, Service={service_name}, Status={display_status}")
            row_idx += 1

        for col_idx in range(1, len(headers) + 1):
            max_length = max(len(str(sheet.cell(row=r, column=col_idx).value or "")) for r in range(1, row_idx))
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2

        ensure_output_directory(output_file)
        workbook.save(output_file)
        print(f"\nReport saved to: {output_file}")

    except Exception as e:
        print(f"Error in report generation: {e}")
        raise

def main():
    # os.system(f'python {base_log_dir}/output_with_raw.py')
    # parser = argparse.ArgumentParser(description="Generate UDS compliance report")
    script_path = os.path.join(base_log_dir, "output_with_raw.py")
    try:
        subprocess.run(
            [sys.executable, script_path],
            check=True,
        )
    except subprocess.CalledProcessError as e:
        print(f"output_with_raw.py failed with return code {e.returncode}")
        sys.exit(e.returncode)

    parser = argparse.ArgumentParser(description="Generate UDS compliance report")
    parser.add_argument("--srd-file", default=SRD_path, help="SRD Excel file path")
    parser.add_argument("--log-file", default=UDS_path, help="Log Excel file path")
    parser.add_argument("--output-file", default=OUTPUT_path, help="Output Excel report")
    parser.add_argument("--extracted-srd-file", default=EXTRACTED_SRD_path, help="Extracted SRD data path")
    parser.add_argument("--srd-sheets", default="DID,NM,Functional Identifiers", help="Comma-separated SRD sheet names")
    parser.add_argument("--log-sheet", default=None, help="Log sheet name")
    args = parser.parse_args()

    try:
        srd_file = validate_file_path(args.srd_file, "SRD file")
        log_file = validate_file_path(args.log_file, "Log file")

        print("\nExtracting SRD services...")
        srd_services, srd_original, srd_details = extract_services_from_srd(srd_file, args.extracted_srd_file, args.srd_sheets.split(","))

        print("\nExtracting log data...")
        log_data, log_original_names, log_dids, log_groups = extract_log_data(log_file, args.log_sheet)

        print("\nGenerating report...")
        compare_and_generate_report(srd_services, srd_original, srd_details, log_data, log_original_names, log_dids, log_groups, args.output_file)

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()