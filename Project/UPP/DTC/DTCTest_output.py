import os
import can
import isotp
from udsoncan.connections import PythonIsoTpConnection
from udsoncan.client import Client
from udsoncan import configs
import pandas as pd
from tabulate import tabulate
from datetime import datetime
import pytz
from openpyxl.styles import Alignment
from openpyxl import load_workbook

# CAN interface setup
can_interface = 'pcan'
channel = 'PCAN_USBBUS2'
tx_id = 0x7D0
rx_id = 0x7D8

# ISO-TP addressing
tp_addr = isotp.Address(isotp.AddressingMode.Normal_11bits, txid=tx_id, rxid=rx_id)
bus = can.Bus(channel=channel, interface=can_interface, bitrate=500000)
stack = isotp.CanStack(bus=bus, address=tp_addr)
conn = PythonIsoTpConnection(stack)

# Ordered list of status flags (bit 0 to bit 7)
status_flags = [
    "test_failed",                              # Bit 0
    "test_failed_this_operation_cycle",         # Bit 1
    "pending",                                  # Bit 2
    "confirmed",                                # Bit 3
    "test_not_completed_since_last_clear",      # Bit 4
    "test_failed_since_last_clear",             # Bit 5
    "test_not_completed_this_operation_cycle",  # Bit 6
    "warning_indicator_requested"               # Bit 7
]

# Map DTC codes to names
dtc_names = {
"31E04B": "MotorOverTemp","31E000": "MotorLowTemp","31E100": "MotorTempSensor","31E24B": "ControllerOverTemp",
"31E200": "ControllerLowTemp","31E300": "ControllerTempSensor","31E44B": "CoolingPlateOverTemp",
"31E400": "CoolingPlateTempSensor","31E54B": "CpuOverTemp","31E664": "McuTempPlausibility",
"31E719": "Phase_U_OverCurrentPeak","31E700": "Phase_U_Sensor","31E819": "Phase_V_OverCurrentPeak",
"31E800": "Phase_W_Sensor","31E919": "Phase_W_OverCurrentPeak","31E901": "Phase_U_Disconnected",
"31EA01": "Phase_V_Disconnected","31EB01": "Phase_W_Disconnected","31EC01": "ActiveShortCircuitFault",
"31ED00": "GateDrvFeedbackActivePbit","31EE00": "GateDrvFeedbackActiveCbit","31EF00": "GateDrvFeedbackInActivePbit",
"31F000": "GateDrvFeedbackInActiveCbit","31F101": "HV_Pbit","31F219": "BusOverCurrent","31F200": "BusCurrentSensor",
"31F216": "BusLowVoltage","31F217": "BusHighVoltage","31F316": "BusUnderVoltage","31F317": "BusOverVoltage",
"31F364": "BusVoltagePlausibility","31F400": "ActDisFeedBackActivePbit","31F500": "ActDisFeedBackActiveCbit",
"31F600": "ActDisFeedBackInActivePbit","31F700": "ActDisFeedBackInActiveCbit","31F800": "ActiveDischargeTimeout",
"31F916": "GateDriverVoltage","31F900": "GateDriversFault","31FA16": "KL30_UnderVoltage","31FA17": "KL30_OverVoltage",
"31FA31": "KL30_Los","31FB00": "PositionSensor","31FC00": "MotorSpeedFault","31FC07": "PRJ_MOTOR_OVERSPEED",
"31FB07": "MotorStall","31FD00": "IoExpenderConfig","31FE00": "CanTransceiver","31FF00": "MemoryRead",
"000000": "CurrentFactory","34F71C": "5V_Sensor","34F800": "CanNoCom","D05403": "PRJ_VCU3_100_TIMEOUT",
"D05407": "PRJ_VCU5_500_TIMEOUT","D05400": "PRJ_VCU8_10_TIMEOUT","D15400": "PRJ_VCU_NODE_ABSENT",
"D05409": "PRJ_VCU14_20_TIMEOUT","D05402": "PRJ_VCU17_500_TIMEOUT","D12600": "PRJ_BMS_NODE_ABSENT",
"D22101": "CanErrorActive","D22102": "CanErrorPassive","D22100": "CanBusOff","E71400": "PRJ_VCU3_100_MISMATCH",
"E40500": "PRJ_VCU5_500_MISMATCH","E71802": "PRJ_VCU8_10_MISMATCH","E75100": "PRJ_VCU9_10_MISMATCH",
"E75200": "PRJ_VCU14_20_MISMATCH","E71200": "PRJ_VCU17_500_MISMATCH","E71401": "PRJ_VCU3_100_ALIVE_CNT",
"E71501": "PRJ_VCU9_10_ALIVE_CNT","E71403": "PRJ_VCU8_10_ALIVE_CNT","E71201": "PRJ_VCU17_500_ALIVE_CNT",
"34F901": "PerfDeratingBusVoltage","34F998": "PerfDeratingMotorTemp","34FA01": "PerfDeratingControllerTemp",
"36A100": "MCU_I2T","D32100": "PRJ_VIN_MISMATCH","34FB46": "ResolverCalibrationRequired","E84B00": "PRJ_FEATURE_CODING",
"D54900": "PRJ_VARIANT_CODING","34FD00": "FusaPbit","34FE00": "FusaCbit","31DF83": "PRJ_CRASH_DETECTION",
"36D815": "PRJ_EMERGENCY_DETECTION"
}

GREEN_CHECK = "\033[1;92m✓\033[0m"  # console (ANSI colored)
PLAIN_CHECK = "✓"                   # Excel (plain)

try:
    with Client(conn, request_timeout=1, config=configs.default_client_config) as client:
        print("Sending ReadDTCInformation (subfunction 0x02 with mask 0xFF)...")
        response = client.read_dtc_information(subfunction=0x02, status_mask=0xFF)

        print(f"\nNumber of DTCs: {response.service_data.dtc_count}")

        # ---- Build rows for console (with ANSI color) ----
        rows_console = []
        # Also build rows for excel (plain checkmark) in parallel
        rows_excel = []

        for dtc in response.service_data.dtcs:
            dtc_id = f"{dtc.id:06X}"
            fault_name = dtc_names.get(dtc_id, "(Unknown)")

            row_console = {"DTC Code": dtc_id, "Fault Name": fault_name}
            row_excel   = {"DTC Code": dtc_id, "Fault Name": fault_name}

            status = dtc.status
            # Use the exact same column order as your console (reversed flags)
            for flag in reversed(status_flags):
                val = bool(getattr(status, flag))
                row_console[flag] = GREEN_CHECK if val else " "
                row_excel[flag]   = PLAIN_CHECK if val else ""
            rows_console.append(row_console)
            rows_excel.append(row_excel)

        df_console = pd.DataFrame(rows_console).astype(str)
        df_excel   = pd.DataFrame(rows_excel).astype(str)

        # ---- Console table (as before) ----
        colalign = []
        for col in df_console.columns:
            colalign.append("left" if col in ["DTC Code", "Fault Name"] else "center")
        print(tabulate(df_console, headers="keys", tablefmt="fancy_grid", showindex=False, colalign=colalign))

        # ---- Save to Excel (timestamped name) ----
        tz = pytz.timezone("Israel")
        stamp = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
        out_dir = os.path.join("C:\\", "Users", os.environ.get("USERNAME", ""), "Documents")
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f"dtc_{stamp}.xlsx")

        with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
            df_excel.to_excel(xw, index=False, sheet_name="DTCs")
            ws = xw.book["DTCs"]

            # widths
            ws.column_dimensions['A'].width = 12   # DTC Code
            ws.column_dimensions['B'].width = 28   # Fault Name
            # center all flag columns
            for col_idx in range(3, 3 + len(status_flags)):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = 30  # wide like your console screenshot
                for row_idx in range(1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

            ws.freeze_panes = "A2"  # freeze header

        print(f"\nExcel saved to: {out_path}")

except Exception as e:
    print(f"❌ Error communicating with ECU: {e}")

finally:
    bus.shutdown()
