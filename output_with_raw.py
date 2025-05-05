import os
import glob
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import sys

def get_latest_log_folder(logs_base_path="Logs"):
    """Find the latest folder in Logs directory."""
    log_folders = [f for f in glob.glob(os.path.join(logs_base_path, "*")) if os.path.isdir(f)]
    if not log_folders:
        raise FileNotFoundError("No log folders found in Logs directory.")
    return max(log_folders, key=os.path.getmtime)

def parse_log_line(line, seen_keys=None):
    if seen_keys is None:
        seen_keys = set()

    # Skip irrelevant lines
    if "Processing file:" in line or "[DEBUG]" in line:
        return None

    # Tester Present
    if "[INFO] - Tester Present: ON" in line:
        did = "Tester Present"
        result = "ON"
        status = "Pass"
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # INFO pattern (capture both Converted and Raw Values)
    info_match = re.search(r"\[INFO\] - (.+?)(?:(?:[:,] (?:Converted result|Converted): (.+?))(?:, Raw Values: ([0-9A-F\s]+))?)?(?:\s+(Pass|Fail))?$", line)
    if info_match:
        did = re.sub(r"\s*Matching Tx and Rx", "", info_match.group(1)).strip()
        did = re.sub(r"\s*Read Data By Identifier\s*$", "", did, flags=re.IGNORECASE).strip()  # Remove Read Data By Identifier
        converted = info_match.group(2).strip() if info_match.group(2) else ""
        raw_values = info_match.group(3).strip() if info_match.group(3) else ""
        status = info_match.group(4).strip() if info_match.group(4) else "Pass"

        # Format result: include both Converted and Raw Values only if Raw Values exists, otherwise use Converted
        if converted and raw_values:
            result = f"Converted: {converted}, Raw: {raw_values}"
        else:
            result = converted if converted else ""

        try:
            if result.replace(".", "").isdigit():
                result = int(result)
            elif result.replace(".", "").replace("-", "").isdigit():
                float_val = float(result)
                if float_val.is_integer():
                    result = int(float_val)
                else:
                    result = float_val
        except:
            pass
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # WARNING pattern
    warning_match = re.search(r"\[WARNING\] - (.+)", line)
    if warning_match:
        did = warning_match.group(1).strip()
        core = re.sub(r"^[0-9A-F]{3,4}\s+", "", did).strip()
        if core in seen_keys:
            return None
        seen_keys.add(core)
        return did, "", "Pass"

    # ERROR: Mismatch type
    mismatch_error = re.search(r"\[ERROR\] - (.+?),\s+(Mismatch Tx and Rx [0-9A-F]+),\s+(Fail|Pass)", line)
    if mismatch_error:
        did = mismatch_error.group(1).strip()
        result = mismatch_error.group(2).strip()
        status = mismatch_error.group(3).strip()
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # ERROR: Negative Response
    neg_resp_error = re.search(r"\[ERROR\] - (.+?)\s+Negative Response: (.+)", line)
    if neg_resp_error:
        did = neg_resp_error.group(1).strip()
        result = f"Negative Response: {neg_resp_error.group(2).strip()}"
        status = "Fail"
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # ERROR: No response from ECU
    no_response = re.search(r"\[ERROR\] - (.+?) No response from ECU detected at (.+)", line)
    if no_response:
        did = no_response.group(1).strip()
        result = f"No response at {no_response.group(2).strip()}"
        status = "Fail"
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # Fallback ERROR
    generic_error = re.search(r"\[ERROR\] - (.+)", line)
    if generic_error:
        did = generic_error.group(1).strip()
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, "", "Fail"

    return None

def generate_excel_report(log_folder):
    folder_name = os.path.basename(log_folder)
    output_excel = os.path.join(log_folder, f"{folder_name}_report.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Log Report"

    # Define headers
    headers = ["File Name", "DID / Sub-service", "Result", "Status"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(size=12, bold=False)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    column_widths = {
        'A': 30,  # File Name
        'B': 50,  # DID / Sub-service
        'C': 80,  # Result (increased to accommodate longer results with Raw Values)
        'D': 10  # Status
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Get all .log files
    log_files = glob.glob(os.path.join(log_folder, "*.log"))
    if not log_files:
        raise FileNotFoundError(f"No .log files found in {log_folder}")

    csv_data = []
    status_counts = {'Pass': 0, 'Fail': 0}
    seen_did_subservices = set()
    for log_file in log_files:
        file_name = os.path.splitext(os.path.basename(log_file))[0]
        print(f"Processing log file: {log_file}")
        with open(log_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                parsed = parse_log_line(line, seen_did_subservices)
                if parsed:
                    did_subservice, result, status = parsed
                    csv_data.append([file_name, did_subservice, result, status])
                    status_counts[status] += 1

    csv_data.sort(key=lambda x: x[0])

    # Write data to Excel
    print(f"Writing {len(csv_data)} rows to Excel")
    for row in csv_data:
        ws.append(row)
        row_num = ws.max_row
        status = row[3]
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            if col == 4:  # Status column
                if status == "Pass":
                    cell.font = Font(size=12, bold=False, color="008000")  # Green
                    cell.alignment = Alignment(horizontal="center")
                elif status == "Fail":
                    cell.font = Font(size=12, bold=True, color="FF0000")  # Red
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.font = Font(size=12, bold=False)
                    cell.alignment = Alignment(horizontal="left")
            else:
                cell.font = Font(size=12, bold=False)
                cell.alignment = Alignment(horizontal="left")

        # Set cell format for Result column
        result_cell = ws.cell(row=row_num, column=3)
        result_cell.alignment = Alignment(horizontal="left")
        if isinstance(row[2], int):
            result_cell.number_format = '0'
        elif isinstance(row[2], float):
            result_cell.number_format = '0.0'
        else:
            result_cell.number_format = '@'

    # Create Charts sheet
    print("Creating Charts sheet")
    try:
        ws_charts = wb.create_sheet("Charts")
        ws_charts.append(["Status", "Count"])
        for status, count in status_counts.items():
            print(f"Status: {status}, Count: {count}")
            ws_charts.append([status, count])
        for cell in ws_charts[1]:
            cell.fill = header_fill
            cell.font = Font(size=12, bold=False)
            cell.alignment = Alignment(horizontal="center")
        ws_charts.column_dimensions['A'].width = 20
        ws_charts.column_dimensions['B'].width = 15

        # Ensure at least one non-zero count to avoid chart rendering issues
        if status_counts['Pass'] == 0 and status_counts['Fail'] == 0:
            print("Warning: No Pass or Fail counts to display in chart. Adding default data.")
            ws_charts.append(["Pass", 1])
            ws_charts.append(["Fail", 0])
            status_counts['Pass'] = 1

        # Generate pie chart with matplotlib
        print("Generating pie chart with matplotlib")
        labels = [
            f"Pass - Count: {status_counts['Pass']}, Pass: {int(round(status_counts['Pass'] / (status_counts['Pass'] + status_counts['Fail']) * 100))}%",
            f"Fail - Count: {status_counts['Fail']}, Fail: {int(round(status_counts['Fail'] / (status_counts['Pass'] + status_counts['Fail']) * 100))}%"
        ]
        sizes = [status_counts['Pass'], status_counts['Fail']]
        colors = ["#008000", "#FF0000"]  # Green for Pass, Red for Fail
        explode = (0.05, 0)  # No explosion of slices

        plt.figure(figsize=(6, 6))
        plt.pie(sizes, explode=explode, labels=labels, colors=colors, autopct=None, startangle=90)
        plt.title(f"Test Results for '{folder_name}'")
        plt.axis('equal')  # Equal aspect ratio ensures pie is circular

        # Save the chart as an image
        chart_path = os.path.join(log_folder, "pie_chart.png")
        plt.savefig(chart_path, bbox_inches='tight')
        plt.close()
        print(f"Pie chart saved as {chart_path}")

        # Embed the image in the Charts sheet at C5
        img = Image(chart_path)
        img.width = 600  # Adjust size as needed
        img.height = 450
        ws_charts.add_image(img, "C5")
        print("Pie chart image added at C5")
    except Exception as e:
        print(f"Error creating chart: {str(e)}")
        # Ensure the Charts sheet is still created even if chart fails
        if "Charts" not in wb.sheetnames:
            wb.create_sheet("Charts")

    # Save Excel file
    try:
        print(f"Saving Excel file: {output_excel}")
        wb.save(output_excel)
        print("Excel file saved successfully")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        sys.exit(1)

    # Clean up the temporary image file
    if os.path.exists(chart_path):
        os.remove(chart_path)
        print(f"Cleaned up temporary file: {chart_path}")

    return output_excel

def main():
    try:
        latest_folder = get_latest_log_folder()
        print(f"Processing log folder: {latest_folder}")
        excel_file = generate_excel_report(latest_folder)
        print(f"Generated Excel report: {excel_file}")
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()