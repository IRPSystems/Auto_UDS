import re
import glob
import os
from datetime import datetime
import pytz
import pandas as pd
import textwrap
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Alignment

# =========================
#  Import dtc_dict
# =========================
try:
    from Condition.dtc_conditions import dtc_dict
    print(f"Debug: Successfully imported dtc_dict. Type: {type(dtc_dict)}")
    # Normalize dtc_dict keys to lowercase
    dtc_dict = {k.lower(): v for k, v in dtc_dict.items()}
    print(f"Debug: Sample dtc_dict keys (normalized): {list(dtc_dict.keys())[:5]}...")
except ImportError as e:
    print(f"Error: Could not import dtc_dict from Condition.dtc_conditions: {e}")
    exit(1)

# =========================
#  Helpers
# =========================
def normalize_repair_actions(text: str) -> str:
    """Reformat repair actions to ensure sequential numbering (1, 2, 3, ...)."""
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return "Refer to diagnostic manual."
    s = str(text)
    lines = s.strip().split('\n')
    steps = [line.strip() for line in lines if re.match(r'^\d+\.\s', line)]
    if not steps:
        return s
    normalized_steps = [f"{i + 1}. {re.sub(r'^\\d+\\.\\s*', '', step)}" for i, step in enumerate(steps)]
    return '\n'.join(normalized_steps)

def _norm_code(s: str) -> str:
    """Normalize a DTC code string to 0x###### lowercase."""
    if s is None:
        return None
    s = str(s).strip().lower()
    s = re.sub(r'[^0-9a-fx]', '', s)
    if not s.startswith('0x'):
        s = '0x' + s
    return s

def _norm_header(h: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(h).lower())

def _extract_version(header: str, lead: str):
    """Extract version tuple from 'SW 3.02.00 State' / 'FW 3.01.11 State'."""
    m = re.search(rf'\b{lead}\s+(\d+)\.(\d+)\.(\d+)\s+state\b', header, flags=re.IGNORECASE)
    if not m:
        return None
    return tuple(int(x) for x in m.groups())

def _meaningful_state_value(s: str) -> str | None:
    """
    Return 'implemented' / 'not implemented' if the cell conveys a meaningful state,
    else None. Blanks/NaN/'-'/TBD/N/A etc. are ignored.
    """
    if s is None:
        return None
    v = str(s).strip().lower()
    if v in {"", "nan", "none", "-", "—", "n/a", "n\\a", "tbd"}:
        return None
    if "not implemented" in v:
        return "not implemented"
    if "implemented" in v:
        # guard against 'not implemented' which is already handled
        return "implemented"
    return None

# =========================
#  Build column ordering for per-row fallback
# =========================
def _build_state_column_order(ICD_df: pd.DataFrame) -> list[tuple[str, tuple[int,int,int], str]]:
    """
    Returns a list of (column_name, version_tuple, family) ordered by preference:
      SW newest -> older SWs -> FW newest -> older FWs -> legacy FW 3.01.11 if present
    family is 'SW' or 'FW'.
    """
    sw_cols = []
    fw_cols = []
    legacy = []

    for c in ICD_df.columns:
        ver = _extract_version(str(c), "SW")
        if ver:
            sw_cols.append((c, ver, "SW"))
            continue
        ver = _extract_version(str(c), "FW")
        if ver:
            # detect legacy 3.01.11 to place at the end (still with FW family)
            if ver == (3, 1, 11):
                legacy.append((c, ver, "FW"))
            else:
                fw_cols.append((c, ver, "FW"))

    # sort newest first
    sw_cols.sort(key=lambda t: t[1], reverse=True)
    fw_cols.sort(key=lambda t: t[1], reverse=True)

    ordered = sw_cols + fw_cols + legacy
    # Debug print a short summary
    short = [f"{fam} {'.'.join(map(str,ver))}: {col}" for col, ver, fam in ordered[:5]]
    print(f"Debug: Column fallback order (first 5): {short}")
    return ordered

# =========================
#  Load FAULT_DETAILS from ICD Excel
# =========================
FAULT_DETAILS = None
ICD_DF = None

try:
    username = os.environ.get('USERNAME', 'unknown')
    if username == 'unknown':
        raise EnvironmentError("USERNAME environment variable not set.")

    excel_file = os.path.join(
        'C:\\', 'Users', username, 'PycharmProjects', 'UDS', 'Documents', 'HD-UP-ICD-243110-DTC.xlsx'
    )
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Excel file {excel_file} not found.")

    # Read the DTCs sheet; your workbook uses header row at line 2 (skip the first row)
    df = pd.read_excel(excel_file, sheet_name="DTCs", header=0, skiprows=1)

    # Clean / standardize header labels
    df.columns = [re.sub(r"\s+", " ", str(c).replace("\n", " ")).strip() for c in df.columns]

    # Drop duplicate column labels if any (keep the first, e.g., for "Repair Actions")
    df = df.loc[:, ~df.columns.duplicated()]

    # Rename ONLY the first N columns (up to "Repair Actions"); extra FW/SW columns kept
    expected_columns = [
        "Item No.", "DTC", "Fault Type Byte (FTB)", "DTC Hex Code", "IRP DTC Name",
        "DTC Description", "DTC Set Conditions", "DTC Maturation Time",
        "DTC Set Threshold", "DTC Heal Conditions", "DTC Dematuration Time",
        "DTC Heal Threshold", "Actions (Error Reaction)", "Severity",
        "Configurable Parameters", "P BIT", "C BIT", "MCU Error Level",
        "TT", "Alert", "Freeze Frame Data", "SW Plan", "Storage/Memory location",
        "Reaction Group Healing", "Repair Actions"
    ]
    if len(df.columns) < len(expected_columns):
        raise ValueError(f"Sheet has only {len(df.columns)} columns; need at least {len(expected_columns)}")
    for i, new_name in enumerate(expected_columns):
        df.rename(columns={df.columns[i]: new_name}, inplace=True)

    print(f"Debug: Loaded DTCs sheet. Columns: {list(df.columns)}")
    print(f"Debug: Number of rows loaded: {len(df)}")

    # Required columns for FAULT_DETAILS
    dtc_col = "DTC Hex Code"
    severity_col = "Severity"
    actions_col = "Actions (Error Reaction)"
    repair_col = "Repair Actions"

    required_columns = {dtc_col, severity_col, actions_col, repair_col}
    if not required_columns.issubset(df.columns):
        raise KeyError(f"Excel does not contain required columns: {required_columns}")

    # Normalize DTC codes and filter invalid
    def normalize_dtc_code(code, row_idx):
        if pd.isna(code) or not isinstance(code, str):
            return None
        code = code.strip().lower()
        code = re.sub(r'[^0-9a-fx]', '', code)
        if not code.startswith('0x'):
            code = '0x' + code
        if re.match(r'^0x[0-9a-f]{6}$', code):
            return code
        return None

    df[dtc_col] = [normalize_dtc_code(code, idx) for idx, code in enumerate(df[dtc_col])]
    df = df[df[dtc_col].notna()]

    FAULT_DETAILS = df.set_index(dtc_col).apply(
        lambda row: {
            "Severity": str(row.get(severity_col, "Unknown")),
            "Actions": str(row.get(actions_col, "None")) if not pd.isna(row.get(actions_col)) else "None",
            "Repair": normalize_repair_actions(row.get(repair_col))
        }, axis=1
    ).to_dict()

    FAULT_DETAILS = {k.lower(): v for k, v in FAULT_DETAILS.items()}
    ICD_DF = df
    print(f"Debug: Loaded FAULT_DETAILS from Excel. Sample keys: {list(FAULT_DETAILS.keys())[:5]}")
    print(f"Debug: Total DTCs loaded from Excel: {len(FAULT_DETAILS)}")

except Exception as e:
    print(f"Error: Failed to load FAULT_DETAILS from Excel: {e}. Exiting.")
    exit(1)

# =========================
#  Parse UDS log
# =========================
def parse_uds_log(log_content: str) -> List[Tuple[str, str]]:
    """Parse UDS log and return [(DTC Hex Code, Status Byte)]."""
    dtcs = []
    response_match = re.search(r"Rx\) Read DTC Information\s*:\s*0x0A 0xFF\s*((?:0x[0-9A-Fa-f]{2}\s*)*)", log_content)
    if not response_match:
        print("Error: Could not find Read DTC Information response.")
        return dtcs
    response_bytes = response_match.group(1).split()
    for i in range(0, len(response_bytes), 4):
        if i + 3 < len(response_bytes):
            dtc_bytes = response_bytes[i:i + 3]
            dtc_hex = f"0x{dtc_bytes[0][2:4]}{dtc_bytes[1][2:4]}{dtc_bytes[2][2:4]}".lower()
            status_byte = response_bytes[i + 3].lower()
            dtcs.append((dtc_hex, status_byte))
    print(f"Debug: Parsed DTCs: {dtcs[:5]}...")
    return dtcs

# =========================
#  Report generator (same layout as your old script)
# =========================
def generate_dtc_report(dtcs: List[Tuple[str, str]], output_excel: str = None, only_faults: bool = False) -> None:
    print("\n=== DTC Report ===")
    print(f"Generated on: {datetime.now(pytz.timezone('Israel')).strftime('%Y-%m-%d %I:%M %p %Z')}")
    if only_faults:
        print("Filter Applied: Showing only DTCs with status byte 0x27 (faults)")
    print("=============================")

    active_faults = 0
    report_data = []
    wrap_width = 80

    for dtc_hex, status_byte in dtcs:
        if dtc_hex == "0x000000":
            continue
        if only_faults and status_byte != "0x27":
            continue

        dtc_name = dtc_dict.get(dtc_hex.lower(), "Unknown DTC")
        if dtc_name == "Unknown DTC":
            print(f"Debug: DTC {dtc_hex} not found in dtc_dict.")

        details = FAULT_DETAILS.get(dtc_hex.lower(), {
            "Severity": "Unknown",
            "Actions": "Unknown",
            "Repair": "Refer to diagnostic manual."
        })

        is_active = status_byte != "0x00"
        if is_active:
            active_faults += 1

        repair_text = details["Repair"]
        if '\n' in repair_text:
            repair_lines = repair_text.split('\n')
            wrapped_lines = []
            for line in repair_lines:
                wrapped = textwrap.wrap(line, width=wrap_width, subsequent_indent='    ')
                wrapped_lines.extend(wrapped)
            formatted_repair = '\n'.join(wrapped_lines)
        else:
            formatted_repair = '\n'.join(textwrap.wrap(repair_text, width=wrap_width, subsequent_indent='    '))

        entry = {
            "DTC Code": dtc_hex,
            "DTC Name": dtc_name,
            "Status Byte": status_byte,
            "Active": "Yes" if is_active else "No",
            "Severity": details["Severity"],
            "Actions": details["Actions"],
            "Repair Actions": formatted_repair
        }
        report_data.append(entry)

        if is_active:
            print(f"\nDTC: {dtc_hex}")
            print(f"Name: {dtc_name}")
            print(f"Status Byte: {status_byte}")
            print(f"Severity: {details['Severity']}")
            print(f"Actions: {details['Actions']}")
            print(f"Repair Actions:\n{formatted_repair}\n{'-' * wrap_width}")

    valid_count = len([d for d in dtcs if d[0] != '0x000000' and (not only_faults or d[1] == '0x27')])
    print(f"\nSummary: {active_faults} active/pending faults detected out of {valid_count} valid DTCs.")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "DTC Report"
        fieldnames = ["DTC Code", "DTC Name", "Status Byte", "Active", "Severity", "Actions", "Repair Actions"]

        for col_idx, field in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        for row_idx, row in enumerate(report_data, start=2):
            for col_idx, field in enumerate(fieldnames, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row[field]
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 80

        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 100

        wb.save(output_excel)
        print(f"Report saved to {output_excel}")
    except Exception as e:
        print(f"Error saving Excel: {e}")

# =========================
#  ICD vs UDS implementation comparison (per-row fallback)
# =========================
def compare_icd_vs_uds(ICD_df: pd.DataFrame, dtcs: List[Tuple[str, str]], out_excel: str = "dtc_impl_check.xlsx"):
    """
    Compare UDS-reported DTCs against ICD implementation status.
    For each ICD row, use the FIRST meaningful state among:
      SW newest -> older SWs -> FW newest -> older FWs -> legacy FW 3.01.11.
    """
    # sets for UDS
    reported_dtcs = {c.lower() for (c, _) in dtcs if c and c.lower() != "0x000000"}

    # column order for per-row fallback
    order = _build_state_column_order(ICD_df)

    # locate DTC column
    icd_code_col = "DTC Hex Code" if "DTC Hex Code" in ICD_df.columns else None
    if icd_code_col is None:
        for c in ICD_df.columns:
            if _norm_header(c) == "dtchexcode":
                icd_code_col = c
                break
    if icd_code_col is None:
        raise KeyError("Could not find 'DTC Hex Code' column in ICD sheet for comparison.")

    codes_series = ICD_df[icd_code_col].dropna().astype(str).map(_norm_code)
    icd_all_codes = set(codes_series)

    icd_impl_codes = set()
    icd_not_impl_codes = set()

    # track which column provided the decision for visibility/debug
    usage_counter = {}

    # Pre-extract the columns to speed per-row lookups
    extracted_cols = [(col, ICD_df[col].astype(str)) for col, _, _ in order]

    for idx, code_raw in codes_series.items():
        state_found = None
        used_col = None
        for col, series in extracted_cols:
            val = _meaningful_state_value(series.get(idx, None))
            if val:
                state_found = val
                used_col = col
                usage_counter[col] = usage_counter.get(col, 0) + 1
                break
        if state_found == "implemented":
            icd_impl_codes.add(code_raw)
        elif state_found == "not implemented":
            icd_not_impl_codes.add(code_raw)
        # else: no info anywhere — ignore for implemented/not-implemented sets

    missing_implemented = sorted(icd_impl_codes - reported_dtcs)
    reported_but_not_impl = sorted((reported_dtcs & icd_all_codes) & icd_not_impl_codes)
    reported_not_in_icd = sorted(reported_dtcs - icd_all_codes)

    print("\n=== ICD vs UDS Implementation Check (Per-row fallback) ===")
    if order:
        first_cols = ", ".join([f"{fam} {'.'.join(map(str,ver))} [{col}]" for col, ver, fam in order[:3]])
        print(f"Column order (top 3): {first_cols}")
    else:
        print("Note: No SW/FW State columns found; skipping implementation status comparison logic.")

    # Show brief usage of columns that contributed decisions
    if usage_counter:
        print("Debug: Decision column usage (first 5):",
              sorted([(k, v) for k, v in usage_counter.items()], key=lambda x: -x[1])[:5])

    print(f"ICD DTCs total: {len(icd_all_codes)}")
    print(f"UDS reported DTCs: {len(reported_dtcs)}")

    print("\n-- Implemented in ICD but NOT seen in UDS log --")
    if missing_implemented:
        for c in missing_implemented:
            print(" ", c)
    else:
        print(" (none)")

    print("\n-- Reported in UDS but marked NOT IMPLEMENTED in ICD --")
    if reported_but_not_impl:
        for c in reported_but_not_impl:
            print(" ", c)
    else:
        print(" (none)")

    print("\n-- Reported in UDS but NOT FOUND in ICD --")
    if reported_not_in_icd:
        for c in reported_not_in_icd:
            print(" ", c)
    else:
        print(" (none)")

    # Output workbook
    try:
        summary = pd.DataFrame({
            "Metric": [
                "ICD total DTCs",
                "UDS reported DTCs",
                "Implemented but missing in UDS",
                "Reported but Not Implemented in ICD",
                "Reported but Not in ICD"
            ],
            "Count": [
                len(icd_all_codes),
                len(reported_dtcs),
                len(missing_implemented),
                len(reported_but_not_impl),
                len(reported_not_in_icd)
            ]
        })

        df_missing_impl = pd.DataFrame({"DTC Code": missing_implemented})
        df_rep_not_impl = pd.DataFrame({"DTC Code": reported_but_not_impl})
        df_rep_not_in_icd = pd.DataFrame({"DTC Code": reported_not_in_icd})

        with pd.ExcelWriter(out_excel, engine="openpyxl") as xw:
            summary.to_excel(xw, index=False, sheet_name="Summary")
            df_missing_impl.to_excel(xw, index=False, sheet_name="Implemented_but_Missing")
            df_rep_not_impl.to_excel(xw, index=False, sheet_name="Reported_but_NotImplemented")
            df_rep_not_in_icd.to_excel(xw, index=False, sheet_name="Reported_but_NotInICD")
        print(f"Implementation comparison saved to {out_excel}")
    except Exception as e:
        print(f"Error saving implementation comparison Excel: {e}")

# =========================
#  Main
# =========================
def main(only_faults: bool = False):
    folder_path = r"C:\temp3"
    files = glob.glob(os.path.join(folder_path, "*.uds.txt"))
    if not files:
        print("No matching files found.")
        return

    latest_file = max(files, key=os.path.getmtime)
    print(f"Processing file: {latest_file}")

    try:
        with open(latest_file, 'r', encoding="utf-8", errors="ignore") as f:
            uds_log = f.read()
    except Exception as e:
        print(f"Error reading file {latest_file}: {e}")
        return

    dtcs = parse_uds_log(uds_log)

    try:
        compare_icd_vs_uds(ICD_DF, dtcs, out_excel="dtc_impl_check.xlsx")
    except Exception as e:
        print(f"Warning: implementation comparison skipped due to error: {e}")

    generate_dtc_report(dtcs, output_excel="dtc_report.xlsx", only_faults=only_faults)

if __name__ == "__main__":
    main(only_faults=False)
