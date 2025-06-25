import re
import glob
import os
from datetime import datetime
import pytz
import pandas as pd
import textwrap
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Import dtc_dict from Condition.dtc_conditions
try:
    from Condition.dtc_conditions import dtc_dict
    print(f"Debug: Successfully imported dtc_dict. Type: {type(dtc_dict)}")
    # Normalize dtc_dict keys to lowercase
    dtc_dict = {k.lower(): v for k, v in dtc_dict.items()}
    print(f"Debug: Sample dtc_dict keys (normalized): {list(dtc_dict.keys())[:5]}...")
except ImportError as e:
    print(f"Error: Could not import dtc_dict from Condition.dtc_conditions: {e}")
    exit(1)

# Function to normalize repair action numbering
def normalize_repair_actions(text: str) -> str:
    """
    Reformat repair actions to ensure sequential numbering (1, 2, 3, ...).
    Args:
        text: Input repair actions string (e.g., "1. Step A\n2. Step B\n4. Step C\n6. Step D")
    Returns:
        str: Reformatted string with corrected numbering (e.g., "1. Step A\n2. Step B\n3. Step C\n4. Step D")
    """
    if not text or pd.isna(text):
        return "Refer to diagnostic manual."

    # Split into lines, preserving content
    lines = text.strip().split('\n')
    # Extract steps (lines starting with a number followed by a period)
    steps = [line.strip() for line in lines if re.match(r'^\d+\.\s', line)]
    if not steps:
        return text  # Return original if no numbered steps

    # Rebuild with sequential numbering
    normalized_steps = [f"{i + 1}. {re.sub(r'^\d+\.\s*', '', step)}" for i, step in enumerate(steps)]
    return '\n'.join(normalized_steps)

# Load FAULT_DETAILS from Excel
FAULT_DETAILS = None
try:
    # Use %USERNAME% environment variable for portable path
    username = os.environ.get('USERNAME', 'unknown')
    if username == 'unknown':
        raise EnvironmentError("USERNAME environment variable not set.")
    excel_file = os.path.join(
        'C:\\', 'Users', username, 'PycharmProjects', 'UDS', 'Documents', 'HD-UP-ICD-243110-DTC.xlsx'
    )

    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Excel file {excel_file} not found.")

    # Read the Excel file, starting at header row (row 2, 0-based index 1)
    df = pd.read_excel(excel_file, sheet_name="DTCs", header=0, skiprows=1)

    # Define expected column names based on document structure
    expected_columns = [
        "Item No.", "DTC", "Fault Type Byte (FTB)", "DTC Hex Code", "IRP DTC Name",
        "DTC Description", "DTC Set Conditions", "DTC Maturation Time",
        "DTC Set Threshold", "DTC Heal Conditions", "DTC Dematuration Time",
        "DTC Heal Threshold", "Actions (Error Reaction)", "Severity",
        "Configurable Parameters", "P BIT", "C BIT", "MCU Error Level",
        "TT", "Alert", "Freeze Frame Data", "SW Plan", "Storage/Memory location",
        "Reaction Group Healing", "Repair Actions"
    ]

    # Ensure the dataframe has the correct columns
    if len(df.columns) >= len(expected_columns):
        df.columns = expected_columns[:len(df.columns)]
    else:
        print(
            f"Warning: Excel sheet has fewer columns ({len(df.columns)}) than expected ({len(expected_columns)}).")
        df.columns = expected_columns[:len(df.columns)] + [f"Unnamed: {i}" for i in
                                                           range(len(df.columns), len(expected_columns))]

    print(f"Debug: Loaded DTCs sheet. Columns: {list(df.columns)}")
    print(f"Debug: Number of rows loaded: {len(df)}")

    # Define required columns
    dtc_col = "DTC Hex Code"
    severity_col = "Severity"
    actions_col = "Actions (Error Reaction)"
    repair_col = "Repair Actions"

    required_columns = {dtc_col, severity_col, actions_col, repair_col}
    if not required_columns.issubset(df.columns):
        raise KeyError(f"Excel does not contain required columns: {required_columns}")

    # Normalize DTC Hex Code values
    def normalize_dtc_code(code, row_idx):
        if pd.isna(code) or not isinstance(code, str):
            return None
        code = code.strip().lower()
        # Remove any non-hex characters except '0x'
        code = re.sub(r'[^0-9a-fx]', '', code)
        # Ensure it starts with '0x'
        if not code.startswith('0x'):
            code = '0x' + code
        # Validate hex format (6 hex digits)
        if re.match(r'^0x[0-9a-f]{6}$', code):
            return code
        return None

    df[dtc_col] = [normalize_dtc_code(code, idx) for idx, code in enumerate(df[dtc_col])]
    # Filter out rows where DTC Hex Code is invalid
    df = df[df[dtc_col].notna()]
    FAULT_DETAILS = df.set_index(dtc_col).apply(
        lambda row: {
            "Severity": str(row.get(severity_col, "Unknown")),
            "Actions": str(row.get(actions_col, "None")) if not pd.isna(row.get(actions_col)) else "None",
            "Repair": normalize_repair_actions(row.get(repair_col))
        }, axis=1
    ).to_dict()
    # Normalize keys
    FAULT_DETAILS = {k.lower(): v for k, v in FAULT_DETAILS.items()}
    print(f"Debug: Loaded FAULT_DETAILS from Excel. Sample keys: {list(FAULT_DETAILS.keys())[:5]}")
    print(f"Debug: Total DTCs loaded from Excel: {len(FAULT_DETAILS)}")

except Exception as e:
    print(f"Error: Failed to load FAULT_DETAILS from Excel: {e}. Exiting.")
    exit(1)

def parse_uds_log(log_content: str) -> List[Tuple[str, str]]:
    """
    Parse UDS log to extract DTCs and status bytes from Read DTC Information response.
    Returns list of tuples: (DTC Hex Code, Status Byte).
    """
    dtcs = []
    # Find the Read DTC Information response line
    response_match = re.search(r"Rx\) Read DTC Information\s*:\s*0x0A 0xFF\s*((?:0x[0-9A-Fa-f]{2}\s*)*)", log_content)
    if not response_match:
        print("Error: Could not find Read DTC Information response.")
        return dtcs

    # Split response into bytes
    response_bytes = response_match.group(1).split()
    # Group bytes into DTCs (3 bytes) + status (1 byte)
    for i in range(0, len(response_bytes), 4):
        if i + 3 < len(response_bytes):
            # Convert 3 bytes to DTC Hex Code (e.g., 0x31 0xE0 0x4B -> 0x31e04b)
            dtc_bytes = response_bytes[i:i + 3]
            dtc_hex = f"0x{dtc_bytes[0][2:4]}{dtc_bytes[1][2:4]}{dtc_bytes[2][2:4]}".lower()
            status_byte = response_bytes[i + 3]
            dtcs.append((dtc_hex, status_byte))
    print(f"Debug: Parsed DTCs: {dtcs[:5]}...")  # Show first 5 DTCs
    return dtcs

def generate_dtc_report(dtcs: List[Tuple[str, str]], output_excel: str = None) -> None:
    """
    Generate a report mapping DTCs to their details using FAULT_DETAILS.
    Save to an Excel file with headers in row 1 and all cells aligned to top.
    """
    print("\n=== DTC Diagnostic Report ===")
    print(f"Generated on: {datetime.now(pytz.timezone('Israel')).strftime('%Y-%m-%d %I:%M %p %Z')}")
    print("=============================")

    active_faults = 0
    report_data = []
    wrap_width = 80

    for dtc_hex, status_byte in dtcs:
        # Skip invalid DTCs (e.g., 0x000000)
        if dtc_hex == "0x000000":
            continue

        # Get DTC name from dtc_dict
        dtc_name = dtc_dict.get(dtc_hex, "Unknown DTC")
        # Debug: Check if DTC was found
        if dtc_name == "Unknown DTC":
            print(f"Debug: DTC {dtc_hex} not found in dtc_dict.")

        # Get fault details from FAULT_DETAILS
        details = FAULT_DETAILS.get(dtc_hex, {
            "Severity": "Unknown",
            "Actions": "Unknown",
            "Repair": "Refer to diagnostic manual."
        })

        # Check if fault is active/pending (non-zero status)
        is_active = status_byte != "0x00"
        if is_active:
            active_faults += 1

        # Format Repair Actions for console and Excel
        repair_text = details["Repair"]
        if '\n' in repair_text:
            # Split into steps and wrap each step
            repair_lines = repair_text.split('\n')
            wrapped_lines = []
            for line in repair_lines:
                # Wrap each line to the specified width, indenting subsequent lines
                wrapped = textwrap.wrap(line, width=wrap_width, subsequent_indent='    ')
                wrapped_lines.extend(wrapped)
            formatted_repair = '\n'.join(wrapped_lines)
        else:
            # Single-line repair action, wrap if needed
            formatted_repair = '\n'.join(textwrap.wrap(repair_text, width=wrap_width, subsequent_indent='    '))

        # Prepare report entry
        entry = {
            "DTC Code": dtc_hex,
            "DTC Name": dtc_name,
            "Status Byte": status_byte,
            "Active": "Yes" if is_active else "No",
            "Severity": details["Severity"],
            "Actions": details["Actions"],
            "Repair Actions": formatted_repair
        }
        report_data.append(entry)

        # Print active/pending faults with formatted output
        if is_active:
            print(f"\nDTC: {dtc_hex}")
            print(f"Name: {dtc_name}")
            print(f"Status Byte: {status_byte}")
            print(f"Severity: {details['Severity']}")
            print(f"Actions: {details['Actions']}")
            print(f"Repair Actions:\n{formatted_repair}\n{'-' * wrap_width}")

    print(
        f"\nSummary: {active_faults} active/pending faults detected out of {len([d for d in dtcs if d[0] != '0x000000'])} valid DTCs.")

    # Save to Excel
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "DTC Report"
        fieldnames = ["DTC Code", "DTC Name", "Status Byte", "Active", "Severity", "Actions", "Repair Actions"]

        # Write headers to row 1
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Write data starting at row 2
        for row_idx, row in enumerate(report_data, start=2):
            for col_idx, field in enumerate(fieldnames, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row[field]
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Set column widths
        ws.column_dimensions['A'].width = 15  # DTC Code
        ws.column_dimensions['B'].width = 30  # DTC Name
        ws.column_dimensions['C'].width = 12  # Status Byte
        ws.column_dimensions['D'].width = 10  # Active
        ws.column_dimensions['E'].width = 12  # Severity
        ws.column_dimensions['F'].width = 20  # Actions
        ws.column_dimensions['G'].width = 80  # Repair Actions (~80 characters)

        # Set row heights for data rows
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 100  # ~5 lines at 20 points each

        # Ensure no extra rows
        ws.sheet_view.view = None  # Reset any view settings
        wb.save(output_excel)
        print(f"Report saved to {output_excel}")
    except Exception as e:
        print(f"Error saving Excel: {e}")

def main():
    # Path to temp3 folder (not user-specific, kept as is unless specified)
    folder_path = r"C:\temp3"
    files = glob.glob(os.path.join(folder_path, "*.uds.txt"))

    if not files:
        print("No matching files found.")
        return

    # Get the latest file based on modification time
    latest_file = max(files, key=os.path.getmtime)
    print(f"Processing file: {latest_file}")

    # Read the UDS log file
    try:
        with open(latest_file, 'r') as f:
            uds_log = f.read()
    except Exception as e:
        print(f"Error reading file {latest_file}: {e}")
        return

    # Parse DTCs from log
    dtcs = parse_uds_log(uds_log)

    # Generate report and save to Excel
    generate_dtc_report(dtcs, output_excel="dtc_report.xlsx")

if __name__ == "__main__":
    main()