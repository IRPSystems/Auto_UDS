import argparse
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import sys
import re

# Define file paths here (modify these paths as needed)
SRD_path = r"C:\Users\ilyar\Downloads\HD-UP-ICD-242601-UDID.xlsx"  # Path to the SRD Excel file
UDS_path = r"C:\Users\ilyar\PycharmProjects\UDS\Logs\02.05.24\02.05.24_report.xlsx"  # Path to the UDS/log Excel file
OUTPUT_path = os.path.join(os.path.dirname(UDS_path), "implementation_report.xlsx")  # Output file for the final report
EXTRACTED_SRD_path = os.path.join(os.path.dirname(UDS_path),
                                  "extracted_srd_data.xlsx")  # Output file for extracted SRD data

# Define the sheet name for the UDS/log file (modify if needed)
UDS_SHEET_NAME = None  # Set to None to use the first sheet, or specify a name like "Report"

# Define services to skip (these will not appear in the output Excel report)
SKIP_Service = {
    "Freeze Frame Content", "Snapshot data", "Extended Snapshot data",
    "CAN WAKEUP Configuration", "Description", "NM Restart Dly Time After Cranking","BMS5_10  DLC Mismatch Threshold","BMS5_10  CRC Mismatch Threshold","BMS5_10  Parity Mismatch Threshold","BMS5_10  DLC Mismatch Healing Threshold",
    "BMS5_10  CRC Mismatch Healing Threshold","BMS5_10  Parity Mismatch Healing Threshold","BMS5_10  Alive Counter Threshold","BMS5_10  Alive Counter Healing Threshold","BMS6_10  DLC Mismatch Threshold","BMS6_10  CRC Mismatch Threshold","BMS6_10  Parity Mismatch Threshold",
     "BMS6_10  DLC Mismatch Healing Threshold","BMS6_10  CRC Mismatch Healing Threshold","BMS6_10  Parity Mismatch Healing Threshold","BMS6_10  Alive Counter Threshold","BMS6_10  Alive Counter Healing Threshold",
    "VCU_Active DLC Mismatch Threshold","VCU_Active Parity Mismatch Threshold","VCU_Active DLC Mismatch Healing Threshold","VCU_Active Parity Mismatch Healing Threshold","VCU5_100 Timeout","VCU_ACTIVE Timeout","VCU_ACTIVE Healing Time","BMS5_10 Timeout","BMS5_10 Healing Time",
   "Network Management Enable","VCU_100 Timeout Healing Time","VCU_100 Parity Mismatch Threshold","VCU_100 DLC Mismatch Healing Threshold","VCU_100 Parity Mismatch Healing Threshold","BAT OverCurrent Fault Charge Current Declaration Settling Time","BAT Under Voltage Fault Declaration Threshold",

}

def validate_file_path(file_path, file_description):
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"{file_description} not found at: {file_path}")
    return file_path


def ensure_output_directory(output_file):
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")


def normalize_service_name(service):
    if not service or not isinstance(service, str):
        return None
    # Remove extra spaces, special characters, and normalize
    service = re.sub(r'\s+', ' ', service.strip())  # Replace multiple spaces with single space
    service = re.sub(r'[^\w\s]', '', service)  # Remove special characters except spaces
    return service.lower()


def strip_prefix(service):
    if not service or not isinstance(service, str):
        return None, None
    # Match a prefix (hex or numeric) followed by a space, e.g., "F18B " or "0100 "
    match = re.match(r'^([0-9A-Fa-f]+)\s+(.*)$', service)
    if match:
        did = match.group(1).strip()  # Extract the DID (e.g., "F18B")
        service_name = match.group(2).strip()  # Extract the service name (e.g., "ECU Manufacturing Date")
        return did, service_name
    return None, service.strip()


def extract_services_from_srd(file_path, extracted_srd_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    services = set()
    services_with_original = {}  # Map normalized to original for reporting
    services_with_details = []  # List to store [group, lid, description, sheet_name]

    # Normalize the SKIP_Service entries for comparison
    normalized_skip_services = {normalize_service_name(s) for s in SKIP_Service if normalize_service_name(s)}

    # Print available sheet names for debugging
    print(f"Available sheets in {file_path}: {workbook.sheetnames}")

    # Sheets to process - Added case-insensitive options for "DID"
    sheets_to_process = ["Functional Identifiers", "NM Identifiers"]
    did_variants = ["DID", "Did", "did"]
    for variant in did_variants:
        if variant in workbook.sheetnames:
            sheets_to_process.append(variant)
            break

    # Create a new workbook for extracted SRD data
    extracted_workbook = openpyxl.Workbook()
    extracted_sheet = extracted_workbook.active
    extracted_sheet.title = "Extracted SRD Data"

    # Write headers to extracted file
    headers = ["Sheet Name", "Group / Function", "LID", "Description"]
    for col_idx, header in enumerate(headers, 1):
        cell = extracted_sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)

    extracted_row_idx = 2

    for sheet_name in sheets_to_process:
        if sheet_name not in workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in {file_path}. Skipping.")
            continue

        sheet = workbook[sheet_name]
        print(f"Processing sheet: {sheet_name}")
        row_count = 0
        for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
            row_count += 1
            group = row[0]  # Column A ("Group / Function")
            lid = row[1]  # Column B ("LID")
            service = row[2]  # Column C ("Description" or "Parameter / Parameter Group")

            if service and isinstance(service, str):
                normalized_service = normalize_service_name(service)
                if normalized_service:
                    if normalized_service in normalized_skip_services:
                        print(f"Skipping service in {sheet_name}, row {row_count + 1} (in SKIP_Service): {service}")
                        continue
                    services.add(normalized_service)
                    services_with_original[normalized_service] = service  # Store original name
                    services_with_details.append((group, lid, service, sheet_name))

                    # Write to extracted SRD file
                    extracted_sheet.cell(row=extracted_row_idx, column=1).value = sheet_name
                    extracted_sheet.cell(row=extracted_row_idx, column=2).value = group
                    extracted_sheet.cell(row=extracted_row_idx, column=3).value = lid
                    extracted_sheet.cell(row=extracted_row_idx, column=4).value = service
                    extracted_row_idx += 1
                else:
                    print(f"Skipping invalid service in {sheet_name}, row {row_count + 1}: {service}")
            else:
                print(f"Skipping row in {sheet_name}, row {row_count + 1}: Service={service} (group={group}, lid={lid})")

        print(f"Processed {row_count} rows in sheet '{sheet_name}'")

    # Adjust column widths for extracted file
    for col in extracted_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        extracted_sheet.column_dimensions[column].width = adjusted_width

    # Save the extracted SRD data
    ensure_output_directory(extracted_srd_path)
    extracted_workbook.save(extracted_srd_path)
    print(f"Extracted SRD data saved to: {extracted_srd_path}")

    return services, services_with_original, services_with_details


def extract_log_data(log_file_path, sheet_name=None):
    workbook = openpyxl.load_workbook(log_file_path, data_only=True)

    if sheet_name is None:
        if not workbook.sheetnames:
            raise ValueError(f"No sheets found in {log_file_path}.")
        sheet_name = workbook.sheetnames[0]
        print(f"No sheet name specified for log file. Using first sheet: '{sheet_name}'")

    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' does not exist in {log_file_path}. Available sheets: {workbook.sheetnames}")

    sheet = workbook[sheet_name]
    log_data = {}
    log_original_names = {}  # Map normalized to original for debugging
    log_dids = {}  # Map normalized service to its DID
    log_groups = {}  # Map normalized service to its group ("File Name")

    # Normalize the SKIP_Service entries for comparison
    normalized_skip_services = {normalize_service_name(s) for s in SKIP_Service if normalize_service_name(s)}

    row_count = 0
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        row_count += 1
        group = row[0]  # Column A ("File Name")
        service = row[1]  # Column B ("DID / Sub-service")
        status = row[3]  # Column D ("Status")
        # Strip the prefix (e.g., "F18B") and capture the DID
        did, service_without_prefix = strip_prefix(service)
        if service_without_prefix:
            normalized_service = normalize_service_name(service_without_prefix)
            if normalized_service and normalized_service not in normalized_skip_services:
                log_data[normalized_service] = status if status else "Unknown"
                log_original_names[normalized_service] = service  # Store original name (with prefix)
                log_dids[normalized_service] = did if did else ""  # Store the DID (or empty string if no DID)
                log_groups[normalized_service] = group  # Store the group ("File Name")
            else:
                print(f"Skipping service (in SKIP_Service) in {sheet_name}, row {row_count + 1}: {service}")
        else:
            print(f"Skipping row in {sheet_name}, row {row_count + 1}: Service={service}")

    print(f"Processed {row_count} rows in log file sheet '{sheet_name}'")
    return log_data, log_original_names, log_dids, log_groups


def compare_and_generate_report(srd_services, srd_original_names, srd_details, log_data, log_original_names, log_dids, log_groups,
                                output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Implementation Report"

    # Define headers to match the template
    headers = ["Group / Function", "LID/DID", "Description", "Status", "Implementation Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)

    # Define fill colors for statuses
    pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for Pass
    #fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for Fail
    fail_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for Fail
    not_impl_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                fill_type="solid")  # Yellow for Not Implemented

    # Compare and populate the report
    row_idx = 2
    unmatched_log_services = set(log_data.keys())  # For tracking services in log but not in SRD
    implemented_nm_services = []
    total_nm_services = 0

    # First, process unmatched log services from specific groups
    groups_to_include = ["Standard_Identifiers", "Generetic_ECU_Read", "Routine_Control"]  # Ordered as requested
    for group in groups_to_include:
        for normalized_service in sorted(unmatched_log_services):
            if log_groups.get(normalized_service) != group:
                continue  # Skip services not in the current group

            # Extract original service name and DID
            original_service = log_original_names[normalized_service]  # e.g., "F186 Active Diagnostic Session"
            did, service_name = strip_prefix(original_service)  # did = "F186", service_name = "Active Diagnostic Session"
            if not service_name:
                continue  # Skip if service name couldn't be parsed

            status = log_data[normalized_service]

            # Column A: Group / Function
            cell_group = sheet.cell(row=row_idx, column=1)
            # Column B: LID/DID
            cell_lid_did = sheet.cell(row=row_idx, column=2)
            # Column C: Description
            cell_description = sheet.cell(row=row_idx, column=3)
            # Column D: Status
            cell_status = sheet.cell(row=row_idx, column=4)
            # Column E: Implementation Status
            cell_impl_status = sheet.cell(row=row_idx, column=5)

            cell_group.value = group
            # Override LID/DID for Routine_Control
            if group == "Routine_Control":
                cell_lid_did.value = "0x31"
            else:
                cell_lid_did.value = did if did else "N/A"  # Use DID if available, otherwise "N/A"
            cell_description.value = service_name

            cell_status.value = status
            if status.lower() == "pass":
                cell_impl_status.value = "Implemented - Pass"
                cell_impl_status.fill = pass_fill
            elif status.lower() == "fail":
                cell_impl_status.value = "Implemented - Waiver"
                cell_impl_status.fill = fail_fill
            else:
                cell_impl_status.value = f"Implemented - Unknown Status ({status})"
                cell_impl_status.fill = fail_fill

            unmatched_log_services.discard(normalized_service)  # Remove from unmatched since it's now processed
            row_idx += 1

    # Then, process all SRD services
    for group, lid, service, sheet_name in srd_details:
        normalized_service = normalize_service_name(service)
        if not normalized_service or normalized_service not in srd_services:
            continue  # Skip if already filtered out (e.g., in SKIP_Service)

        # Count services from "NM Identifiers"
        if sheet_name == "NM Identifiers":
            total_nm_services += 1

        # Column A: Group / Function
        cell_group = sheet.cell(row=row_idx, column=1)
        # Column B: LID/DID
        cell_lid_did = sheet.cell(row=row_idx, column=2)
        # Column C: Description
        cell_description = sheet.cell(row=row_idx, column=3)
        # Column D: Status
        cell_status = sheet.cell(row=row_idx, column=4)
        # Column E: Implementation Status
        cell_impl_status = sheet.cell(row=row_idx, column=5)

        cell_group.value = group
        cell_description.value = service

        # Override LID based on the Group / Function category
        normalized_group = normalize_service_name(group)
        if normalized_group:
            if "can configuration" in normalized_group:
                assigned_lid = "F1D5"
            elif "true drive" in normalized_group:
                assigned_lid = "0790"
            elif "faults configuration" in normalized_group:
                assigned_lid = "078F"
            elif "routine control" in normalized_group:
                assigned_lid = "0x31"
            else:
                assigned_lid = lid  # Use the original LID if no override applies
        else:
            assigned_lid = lid  # Fallback to original LID if group normalization fails

        if normalized_service in log_data:
            unmatched_log_services.discard(normalized_service)  # Matched, so remove from unmatched
            status = log_data[normalized_service]
            # Set LID/DID: Use DID if available, otherwise use the assigned LID
            did_value = log_dids.get(normalized_service, "")
            cell_lid_did.value = did_value if did_value else assigned_lid
            cell_status.value = status
            if status.lower() == "pass":
                cell_impl_status.value = "Implemented - Pass"
                cell_impl_status.fill = pass_fill
                if sheet_name == "NM Identifiers":
                    implemented_nm_services.append(service)
            elif status.lower() == "fail":
                cell_impl_status.value = "Implemented - Waiver"
                cell_impl_status.fill = fail_fill
                if sheet_name == "NM Identifiers":
                    implemented_nm_services.append(service)
            else:
                cell_impl_status.value = f"Implemented - Unknown Status ({status})"
                cell_impl_status.fill = fail_fill
                if sheet_name == "NM Identifiers":
                    implemented_nm_services.append(service)
        else:
            cell_lid_did.value = assigned_lid if assigned_lid else ""  # No DID since no match in log file
            cell_status.value = ""  # Status is empty for non-implemented services
            cell_impl_status.value = "Not Implemented"
            cell_impl_status.fill = not_impl_fill

        row_idx += 1

    # Adjust column widths for readability
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    # Save the report
    ensure_output_directory(output_file)
    workbook.save(output_file)
    print(f"Implementation report generated: {output_file}")

    # Summary of NM Identifiers implementation
    print(f"\nSummary for 'NM Identifiers' sheet:")
    print(f"Total services in 'NM Identifiers': {total_nm_services}")
    print(f"Implemented services in 'NM Identifiers': {len(implemented_nm_services)}")
    if implemented_nm_services:
        print("Implemented 'NM Identifiers' services:")
        for service in implemented_nm_services:
            print(f"- {service}")
    else:
        print("No services from 'NM Identifiers' were implemented.")

    # Log unmatched services for debugging (excluding the ones we included)
    remaining_unmatched = {s for s in unmatched_log_services if log_groups.get(s) not in groups_to_include}
    if remaining_unmatched:
        print("\nServices in log file but not in SRD (potential mismatches):")
        for normalized_service in sorted(remaining_unmatched):
            original_service = log_original_names.get(normalized_service, normalized_service)
            print(f"- {original_service}")


def main():
    parser = argparse.ArgumentParser(
        description="Compare SRD and log Excel files to generate an implementation report.")
    parser.add_argument("--srd-file", default=SRD_path, help=f"Path to the SRD Excel file (default: {SRD_path})")
    parser.add_argument("--log-file", default=UDS_path, help=f"Path to the log Excel file (default: {UDS_path})")
    parser.add_argument("--output-file", default=OUTPUT_path,
                        help=f"Path to the output Excel report file (default: {OUTPUT_path})")
    parser.add_argument("--extracted-srd-file", default=EXTRACTED_SRD_path,
                        help=f"Path to the extracted SRD data file (default: {EXTRACTED_SRD_path})")
    parser.add_argument("--uds-sheet", default=UDS_SHEET_NAME,
                        help="Sheet name in the UDS/log Excel file (default: first sheet)")

    args = parser.parse_args()

    try:
        # Validate file paths
        srd_file = validate_file_path(args.srd_file, "SRD Excel file (HD-UP-ICD-242601-UDID.xlsx)")
        log_file = validate_file_path(args.log_file, "Log Excel file (02.05.24_report.xlsx)")

        # Extract services from SRD and save to a new file
        print("Extracting services from SRD file...")
        srd_services, srd_original_names, srd_details = extract_services_from_srd(srd_file, args.extracted_srd_file)

        # Extract log data
        print("Extracting log data...")
        log_data, log_original_names, log_dids, log_groups = extract_log_data(log_file, sheet_name=args.uds_sheet)

        # Generate the report
        print("Generating implementation report...")
        compare_and_generate_report(srd_services, srd_original_names, srd_details, log_data, log_original_names,
                                    log_dids, log_groups, args.output_file)

    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()