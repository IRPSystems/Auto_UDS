import os
import glob
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, colors

def get_latest_log_folder(logs_base_path="Logs"):
    """Find the latest folder in Logs directory."""
    log_folders = [f for f in glob.glob(os.path.join(logs_base_path, "*")) if os.path.isdir(f)]
    if not log_folders:
        raise FileNotFoundError("No log folders found in Logs directory.")
    return max(log_folders, key=os.path.getmtime)

def parse_log_line(line, seen_keys=None):
    if seen_keys is None:
        seen_keys = set()

    # Skip irrelevant lines
    if "Processing file:" in line or "[DEBUG]" in line:
        return None

    # Tester Present
    if "[INFO] - Tester Present: ON" in line:
        did = "Tester Present"
        result = "ON"
        status = "Pass"
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # INFO pattern
    info_match = re.search(r"\[INFO\] - (.+?)(?:(?:[:,] (?:Converted result|Converted): (.+?))(?:, Raw Values: [0-9A-F\s]+)?)?(?:\s+(Pass|Fail))?$", line)
    if info_match:
        did = re.sub(r"\s*Matching Tx and Rx", "", info_match.group(1)).strip()
        result = info_match.group(2).strip() if info_match.group(2) else ""
        status = info_match.group(3).strip() if info_match.group(3) else "Pass"
        if status != "Pass":
            return None
        try:
            if result.replace(".", "").isdigit():
                result = int(result)
            elif result.replace(".", "").replace("-", "").isdigit():
                result = float(result)
        except:
            pass
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # WARNING pattern
    warning_match = re.search(r"\[WARNING\] - (.+)", line)
    if warning_match:
        did = warning_match.group(1).strip()
        core = re.sub(r"^[0-9A-F]{3,4}\s+", "", did).strip()
        if core in seen_keys:
            return None
        seen_keys.add(core)
        return did, "", "Pass"

    # ERROR: Mismatch type (main case)
    mismatch_error = re.search(r"\[ERROR\] - (.+?),\s+(Mismatch Tx and Rx [0-9A-F]+),\s+(Fail|Pass)", line)
    if mismatch_error:
        did = mismatch_error.group(1).strip()
        result = mismatch_error.group(2).strip()
        status = mismatch_error.group(3).strip()
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # ERROR: Negative Response
    neg_resp_error = re.search(r"\[ERROR\] - (.+?)\s+Negative Response: (.+)", line)
    if neg_resp_error:
        did = neg_resp_error.group(1).strip()
        result = f"Negative Response: {neg_resp_error.group(2).strip()}"
        status = "Fail"
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # ERROR: No response from ECU
    no_response = re.search(r"\[ERROR\] - (.+?) No response from ECU detected at (.+)", line)
    if no_response:
        did = no_response.group(1).strip()
        result = f"No response at {no_response.group(2).strip()}"
        status = "Fail"
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, result, status

    # Fallback ERROR
    generic_error = re.search(r"\[ERROR\] - (.+)", line)
    if generic_error:
        did = generic_error.group(1).strip()
        if did in seen_keys:
            return None
        seen_keys.add(did)
        return did, "", "Fail"

    return None




def generate_excel_report(log_folder):
    folder_name = os.path.basename(log_folder)
    output_excel = os.path.join(log_folder, f"{folder_name}_report.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Log Report"

    # Define headers
    headers = ["File Name", "DID / Sub-service", "Result", "Status"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(size=12, bold=False)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    column_widths = {
        'A': 30,  # File Name
        'B': 80,  # DID / Sub-service
        'C': 40,  # Result
        'D': 15  # Status
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Get all .log files
    log_files = glob.glob(os.path.join(log_folder, "*.log"))
    if not log_files:
        raise FileNotFoundError(f"No .log files found in {log_folder}")

    csv_data = []
    seen_did_subservices = set()
    for log_file in log_files:
        file_name = os.path.splitext(os.path.basename(log_file))[0]
        with open(log_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                parsed = parse_log_line(line, seen_did_subservices)
                if parsed:
                    did_subservice, result, status = parsed
                    csv_data.append([file_name, did_subservice, result, status])

    csv_data.sort(key=lambda x: x[0])

    # Write data to Excel
    for row in csv_data:
        ws.append(row)
        row_num = ws.max_row
        status = row[3]
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            if col == 4:  # Status column
                if status == "Pass":
                    cell.font = Font(size=12, bold=False, color="008000")  # Green
                elif status == "Fail":
                    cell.font = Font(size=12, bold=True, color="FF0000")  # Red
                else:
                    cell.font = Font(size=12, bold=False)
            else:
                cell.font = Font(size=12, bold=False)

        # Set cell format for Result column
        result_cell = ws.cell(row=row_num, column=3)
        if isinstance(row[2], int):
            result_cell.number_format = '0'
        elif isinstance(row[2], float):
            result_cell.number_format = '0.0'
        else:
            result_cell.number_format = '@'

    # Save Excel file
    wb.save(output_excel)
    return output_excel

def main():
    try:
        latest_folder = get_latest_log_folder()
        print(f"Processing log folder: {latest_folder}")
        excel_file = generate_excel_report(latest_folder)
        print(f"Generated Excel report: {excel_file}")
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()